from __future__ import annotations

import argparse
import ctypes
import io
import json
import math
import os
import re
import shutil
import sys
import urllib.error
import urllib.request
from datetime import date, datetime, time
from dataclasses import dataclass, asdict
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel
from pypdf import PdfReader


LOCAL_PACKAGES = Path(__file__).resolve().parent / ".python_packages"
if LOCAL_PACKAGES.exists() and str(LOCAL_PACKAGES) not in sys.path:
    sys.path.insert(0, str(LOCAL_PACKAGES))


DEEPSEEK_API_BASE = os.environ.get("DEEPSEEK_API_BASE", "https://api.deepseek.com").rstrip("/")
DEEPSEEK_API_URL = os.environ.get("DEEPSEEK_API_URL", f"{DEEPSEEK_API_BASE}/chat/completions")
DEEPSEEK_MODEL = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
AI_WARNINGS: list[str] = []


def clear_ai_warnings() -> None:
    AI_WARNINGS.clear()


def get_ai_warnings() -> list[str]:
    return list(AI_WARNINGS)


def record_ai_warning(message: str) -> None:
    if message not in AI_WARNINGS:
        AI_WARNINGS.append(message)


def deepseek_api_url() -> str:
    api_url = os.environ.get("DEEPSEEK_API_URL", "").strip()
    if api_url:
        return api_url
    api_base = os.environ.get("DEEPSEEK_API_BASE", DEEPSEEK_API_BASE).strip().rstrip("/")
    return f"{api_base}/chat/completions"


def load_env_file(path: Path | None = None) -> None:
    env_path = path or Path(__file__).resolve().parent / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if key and not os.environ.get(key):
            os.environ[key] = value


STOP_WORDS = {
    "для",
    "при",
    "под",
    "или",
    "без",
    "над",
    "мм",
    "м2",
    "м3",
    "шт",
    "уп",
    "упак",
    "рул",
    "кг",
    "л",
    "лист",
    "наружная",
    "наружный",
    "серый",
    "белый",
    "оцинкованный",
    "заказ",
    "цвет",
    "расход",
    "аналог",
}

SERVICE_WORDS = {
    "доставка",
    "доставки",
    "транспортные",
    "транспортная",
    "транспорт",
    "услуга",
    "услуги",
    "разгрузка",
    "подъем",
}

SYNONYM_GROUPS = {
    "саморез": {"саморез", "саморезы", "шуруп", "шурупы", "метиз", "метизы", "крепеж", "крепежный", "винт", "клопы", "гм"},
    "гипсокартон": {"гипсокартон", "гкл", "сапфир"},
    "мембрана": {"мембрана", "ветрозащита", "ветро", "влагозащита", "пароизоляция", "пароизоляционная", "гидро", "изоспан"},
    "утеплитель": {"утеплитель", "теплоизоляция", "изоляция", "вата", "минвата", "каменная", "стеклянного"},
    "профиль": {"профиль", "каркас", "планка", "направляющая", "поперечная", "угловой", "j", "l", "f", "омега", "пи"},
    "плита": {"плита", "панель", "панели", "аквапанель", "акустическая", "потолочная"},
    "клей": {"клей", "клеевая", "клеевой", "штукатурно", "смесь", "смеси", "цементная"},
    "грунтовка": {"грунтовка", "грунт", "тифенгрунт", "ct17", "ст17"},
    "гидроизоляция": {"гидроизоляция", "гидроизоляционный", "флехендихт", "флэхендихт", "обмазочная"},
    "пленка": {"пленка", "пленка", "полиэтиленовая", "укрывочная", "микрон", "мкм"},
    "затирка": {"затирка", "затирочная", "ce40", "се40"},
    "сетка": {"сетка", "стеклосетка", "стеклотканевая", "фасадная"},
    "рейка": {"рейка", "рейки", "каркас", "норма", "norma", "т24", "t24", "планка"},
    "гребенка": {"гребенка", "bts", "bt", "вт"},
    "подвес": {"подвес", "подвесов", "нониус", "европодвес"},
    "соединитель": {"соединитель", "соед", "элем", "элемент"},
    "герметик": {"герметик", "силикон", "силиконовый"},
    "плитка": {"плитка", "керамогранит", "керамическая"},
    "краска": {"краска", "окраска", "акриловая"},
    "штукатурка": {"штукатурка", "штукатурки", "декоративная", "камешковая"},
    "водосток": {"водосток", "водосточная", "сток", "слив", "колено", "труба", "муфта", "хомут"},
    "osb": {"osb", "осп", "osb3", "osb-3"},
    "кирпич": {"кирпич", "кирпичи", "кладочный", "кладочная"},
    "пиломатериал": {"пиломатериал", "доска", "брус", "рейка", "стропила", "лагa", "лаги"},
    "арматура": {"арматура", "арматурный", "а500", "а400", "рифленая"},
}

SYNONYM_BY_TOKEN = {
    token: canonical
    for canonical, variants in SYNONYM_GROUPS.items()
    for token in variants
}

KNOWN_BRANDS = {
    "knauf": "кнауф",
    "кнауф": "кнауф",
    "volma": "волма",
    "волма": "волма",
    "grandline": "grandline",
    "грандлайн": "grandline",
    "изоспан": "изоспан",
    "технониколь": "технониколь",
    "техновент": "техновент",
    "церезит": "церезит",
    "ceresit": "церезит",
    "rockfon": "rockfon",
    "рокфон": "rockfon",
    "albes": "albes",
    "албес": "albes",
}


@dataclass
class RequestItem:
    pos: str
    name: str
    specs: str = ""
    unit: str = ""
    qty: float | None = None


@dataclass
class SupplierItem:
    supplier: str
    source: str
    row_no: str
    name: str
    qty: float | None = None
    unit: str = ""
    price: float | None = None
    total: float | None = None
    delivery: str = ""
    invoice_no: str = ""
    override_qty: float | None = None
    override_unit: str = ""


@dataclass
class Match:
    supplier_item: SupplierItem
    request_pos: str | None
    score: float
    status: str
    reason: str


@dataclass(frozen=True)
class QuantityCheck:
    status: str
    display: str
    converted_qty: float | None = None
    converted_unit: str = ""


STATUS_LABELS = {
    "auto": "автоматически",
    "review": "на проверку",
    "unmatched": "не сопоставлено",
    "manual": "подтверждено на проверке",
    "service": "доставка/услуга",
}


def status_label(status: str) -> str:
    return STATUS_LABELS.get(status, status)


@lru_cache(maxsize=20000)
def fix_mojibake(text: str) -> str:
    if not text or not re.search(r"(Р[°-џ]|\u0098|вЂ)", text):
        return text
    try:
        raw = bytearray()
        for char in text:
            code = ord(char)
            if code <= 255:
                raw.append(code)
            else:
                raw.extend(char.encode("cp1251"))
        fixed = raw.decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError, ValueError):
        return text

    bad_before = len(re.findall(r"(Р[°-џ]|\u0098|вЂ)", text))
    bad_after = len(re.findall(r"(Р[°-џ]|\u0098|вЂ)", fixed))
    if bad_after < bad_before:
        return fixed
    return text


@lru_cache(maxsize=40000)
def clean_text(value) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    for _ in range(2):
        fixed = fix_mojibake(text)
        if fixed == text:
            break
        text = fixed
    return text


def parse_number(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime):
        return float(to_excel(value))
    if isinstance(value, date):
        return float(to_excel(datetime.combine(value, time())))
    text = clean_text(value)
    text = text.replace(" ", "").replace("\u202f", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_quantity(value, price: float | None = None, total: float | None = None) -> float | None:
    qty = parse_number(value)
    if qty is None:
        return None
    if not price or not total or price <= 0 or total <= 0:
        return qty

    expected = total / price
    if expected <= 0 or expected >= 1_000_000:
        return qty

    smaller = max(min(abs(qty), abs(expected)), 0.000001)
    ratio = max(abs(qty), abs(expected)) / smaller
    amount_error = abs((qty * price) - total)

    if (qty >= 1_000_000 and ratio >= 10) or (ratio >= 50 and amount_error > max(total * 0.2, 100)):
        return round(expected, 4)
    return qty


def money_to_float(text: str) -> float | None:
    return parse_number(text)


@lru_cache(maxsize=40000)
def normalize(text: str) -> str:
    text = clean_text(text).lower().replace("ё", "е")
    text = text.replace("x", "х").replace("*", "х")
    text = re.sub(r"(?<=\d),(?=\d)", ".", text)
    text = re.sub(r"[^0-9a-zа-я.х]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


@lru_cache(maxsize=40000)
def tokens(text: str) -> frozenset[str]:
    base_tokens = {
        token
        for token in normalize(text).split()
        if len(token) > 1 and token not in STOP_WORDS
    }
    expanded = set(base_tokens)
    for token in base_tokens:
        canonical = SYNONYM_BY_TOKEN.get(token)
        if canonical:
            expanded.add(canonical)
    return frozenset(expanded)


@lru_cache(maxsize=40000)
def is_service_text(text: str) -> bool:
    text_tokens = tokens(text)
    normalized = normalize(text)
    if text_tokens & SERVICE_WORDS:
        return True
    return any(marker in normalized for marker in ["доставка", "транспортные услуги", "транспортная услуга"])


def is_service_item(item: SupplierItem) -> bool:
    return is_service_text(item.name)


@lru_cache(maxsize=40000)
def item_categories(text: str) -> frozenset[str]:
    return frozenset(SYNONYM_BY_TOKEN[token] for token in tokens(text) if token in SYNONYM_BY_TOKEN)


@lru_cache(maxsize=40000)
def product_type_tags(text: str) -> frozenset[str]:
    source = normalize(text)
    tags: set[str] = set()

    patterns = [
        ("thermostatic_head", (r"\bтермоголовк", r"термостатическ\w*\s+головк", r"головк\w*\s+термостат")),
        ("radiator", (r"\bрадиатор",)),
        ("thermostatic_valve", (r"клапан\w*\s+термостат", r"термостат\w*\s+клапан")),
        ("radiator_valve", (r"клапан\w*\s+радиатор", r"настроечн\w*\s+клапан")),
        ("vacuum_valve", (r"клапан\w*\s+вакуум", r"\bаэратор")),
        ("ball_valve", (r"кран\w*\s+шаров", r"шаров\w*\s+кран")),
        ("air_vent", (r"воздухоотвод",)),
        ("pipe", (r"\bтруб",)),
        ("elbow", (r"\bотвод", r"\bугол")),
        ("tee", (r"\bтройник",)),
        ("transition", (r"\bпереход",)),
        ("cross", (r"\bкрестовин",)),
        ("plug", (r"\bзаглушк",)),
        ("coupling", (r"\bмуфт",)),
        ("clamp", (r"\bхомут",)),
        ("bolt", (r"\bболт",)),
        ("screw", (r"\bвинт", r"\bсаморез", r"\bшуруп")),
        ("nut", (r"\bгайк",)),
        ("sealant", (r"\bгерметик", r"\bсиликон")),
        ("primer", (r"\bгрунтовк", r"\bгрунт\b")),
    ]
    for tag, tag_patterns in patterns:
        if any(re.search(pattern, source) for pattern in tag_patterns):
            tags.add(tag)

    if not tags and re.search(r"\bклапан", source):
        tags.add("valve")
    if not tags and re.search(r"\bкран", source):
        tags.add("tap")
    return frozenset(tags)


def product_types_incompatible(request_name: str, supplier_name: str) -> bool:
    req_tags = product_type_tags(request_name)
    sup_tags = product_type_tags(supplier_name)
    return bool(req_tags and sup_tags and req_tags.isdisjoint(sup_tags))


def normalized_unit(unit: str) -> str:
    raw = clean_text(unit).lower().replace("²", "2").replace("^2", "2").replace("³", "3").replace("^3", "3")
    raw = raw.replace(" ", "").replace(".", "")
    if raw in {"м2", "квм", "квм2"}:
        return "m2"
    if raw in {"м3", "кубм", "куб"}:
        return "m3"
    if raw in {"м", "мп", "пм"}:
        return "m"
    if raw in {"шт", "штук"}:
        return "pcs"
    if raw in {"лист", "листы"}:
        return "sheet"
    if raw in {"рул", "рулон", "рулоны"}:
        return "roll"
    if raw in {"уп", "упак", "упаковка"}:
        return "pack"
    if raw in {"кг"}:
        return "kg"
    if raw in {"т", "тн", "тонна", "тонны"}:
        return "ton"
    unit_norm = normalize(unit)
    if unit_norm in {"м2", "м.2", "кв.м", "кв", "м²"}:
        return "m2"
    if unit_norm in {"м3", "м.3", "куб.м", "куб"}:
        return "m3"
    if unit_norm in {"м", "м.п", "п.м", "мп"}:
        return "m"
    if unit_norm in {"шт", "штук"}:
        return "pcs"
    if unit_norm in {"лист", "листы"}:
        return "sheet"
    if unit_norm in {"рул", "рулон", "рулоны"}:
        return "roll"
    if unit_norm in {"уп", "упак", "упаковка"}:
        return "pack"
    if unit_norm in {"кг"}:
        return "kg"
    if unit_norm in {"т", "тн", "тонна", "тонны"}:
        return "ton"
    return unit_norm


@lru_cache(maxsize=40000)
def area_m2_from_text(text: str) -> float | None:
    normalized = normalize(text)
    area_patterns = [
        r"(?:s|площадь)\s*=?\s*(\d+(?:\.\d+)?)\s*(?:м2|м²|m2|m²|кв\.?м)",
        r"(\d+(?:\.\d+)?)\s*(?:м2|м²|m2|m²|кв\.?м)",
    ]
    for pattern in area_patterns:
        match = re.search(pattern, normalized)
        if match:
            value = parse_number(match.group(1))
            if value:
                return value
    dim = re.search(r"(\d{3,4})\s*х\s*(\d{3,4})", normalized)
    if dim:
        first = parse_number(dim.group(1))
        second = parse_number(dim.group(2))
        if first and second:
            return round(first * second / 1_000_000, 4)
    return None


def _dimensions_m_from_text(text: str) -> tuple[float, ...]:
    prepared = clean_text(text).lower()
    prepared = prepared.replace("×", "x").replace("*", "x").replace("х", "x").replace(",", ".")
    match = re.search(r"(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))?", prepared)
    if not match:
        return ()
    values = [parse_number(value) for value in match.groups() if value]
    if not values or any(value is None for value in values):
        return ()
    numbers = [float(value) for value in values if value is not None]
    if any(number > 50 for number in numbers):
        return tuple(number / 1000 for number in numbers)
    return tuple(numbers)


@lru_cache(maxsize=40000)
def dimension_area_m2_from_text(text: str) -> float | None:
    dims = _dimensions_m_from_text(text)
    if len(dims) >= 2:
        return round(dims[0] * dims[1], 4)
    return None


@lru_cache(maxsize=40000)
def volume_m3_from_text(text: str) -> float | None:
    normalized = normalize(text)
    volume_patterns = [
        r"(\d+(?:\.\d+)?)\s*(?:м3|м³|m3|m³|куб\.?м)",
    ]
    for pattern in volume_patterns:
        match = re.search(pattern, normalized)
        if match:
            value = parse_number(match.group(1))
            if value:
                return value
    dims = _dimensions_m_from_text(text)
    if len(dims) >= 3:
        return round(dims[0] * dims[1] * dims[2], 6)
    return None


@lru_cache(maxsize=40000)
def length_m_from_text(text: str) -> float | None:
    source = clean_text(text).lower().replace(",", ".")
    patterns = [
        r"(?:l|длина)\s*=?\s*(\d+(?:\.\d+)?)\s*(?:м|m)\b",
        r"\b(\d+(?:\.\d+)?)\s*(?:м|m)\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, source):
            value = parse_number(match.group(1))
            if value and 0.05 <= value <= 20:
                return value
    return None


@lru_cache(maxsize=40000)
def long_dimension_m_from_text(text: str) -> float | None:
    source = clean_text(text).lower().replace(",", ".")
    patterns = [
        r"(?:l|длина)\s*=?\s*(\d+(?:\.\d+)?)\s*мм",
        r"\b(\d{3,5}(?:\.\d+)?)\s*мм\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, source):
            value = parse_number(match.group(1))
            if value and 100 <= value <= 20000:
                return value / 1000
    dims = _dimensions_m_from_text(text)
    if dims:
        longest = max(dims)
        if 0.1 <= longest <= 20:
            return longest
    return None


@lru_cache(maxsize=40000)
def package_count_from_text(text: str) -> float | None:
    source = clean_text(text).lower().replace(",", ".")
    patterns = [
        r"(\d+(?:\.\d+)?)\s*шт\s*/\s*(?:уп|упак|пачк|кор)",
        r"(?:уп|упак|пачк|кор)[^\d]{0,20}(\d+(?:\.\d+)?)\s*шт",
        r"(\d+(?:\.\d+)?)\s*шт\s*/",
        r"(\d+(?:\.\d+)?)\s*шт\.?\s*[,) ]",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, source):
            value = parse_number(match.group(1))
            if value and 1 < value <= 10000:
                return value
    return None


@lru_cache(maxsize=40000)
def package_kg_from_text(text: str) -> float | None:
    source = clean_text(text).lower().replace(",", ".")
    patterns = [
        r"(\d+(?:\.\d+)?)\s*кг\s*/\s*(?:уп|упак|пачк|меш)",
        r"(?:уп|упак|пачк|меш)[^\d]{0,20}(\d+(?:\.\d+)?)\s*кг",
        r"\b(\d+(?:\.\d+)?)\s*кг\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, source):
            value = parse_number(match.group(1))
            if value and 0.05 <= value <= 1000:
                return value
    return None


@lru_cache(maxsize=40000)
def package_linear_m_from_text(text: str) -> float | None:
    source = clean_text(text).lower().replace(",", ".")
    patterns = [
        r"(\d+(?:\.\d+)?)\s*(?:м\.?п\.?|п\.?м\.?|мп)\s*/\s*(?:уп|упак|пачк|кор)",
        r"(?:уп|упак|пачк|кор)[^\d]{0,20}(\d+(?:\.\d+)?)\s*(?:м\.?п\.?|п\.?м\.?|мп)",
        r"/\s*(\d+(?:\.\d+)?)\s*(?:м\.?п\.?|п\.?м\.?|мп)",
        r"(\d+(?:\.\d+)?)\s*(?:м\.?п\.?|п\.?м\.?|мп)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, source):
            value = parse_number(match.group(1))
            if value and 0.05 <= value <= 10000:
                return value
    return None


@lru_cache(maxsize=40000)
def is_linear_profile_text(text: str) -> bool:
    normalized = normalize(text)
    markers = {
        "профиль",
        "профил",
        "рейка",
        "рейки",
        "каркас",
        "планка",
        "подвес",
        "уголок",
        "направляющая",
        "доска",
        "брус",
        "пиломатериал",
        "вагонка",
    }
    return any(marker in normalized for marker in markers)


@lru_cache(maxsize=40000)
def is_lumber_text(text: str) -> bool:
    normalized = normalize(text)
    markers = {
        "пиломатериал",
        "доска",
        "доски",
        "брус",
        "брусок",
        "рейка",
        "стропила",
        "лага",
        "лаги",
        "вагонка",
    }
    return any(marker in normalized for marker in markers)


@lru_cache(maxsize=40000)
def lumber_piece_volume_m3_from_text(text: str) -> float | None:
    dims = _dimensions_m_from_text(text)
    if len(dims) >= 3:
        return round(dims[0] * dims[1] * dims[2], 6)
    return volume_m3_from_text(text)


@lru_cache(maxsize=40000)
def lumber_cross_section_m2_from_text(text: str) -> float | None:
    dims = _dimensions_m_from_text(text)
    if len(dims) >= 3:
        short = sorted(dims)[:2]
        return round(short[0] * short[1], 6)
    if len(dims) == 2:
        return round(dims[0] * dims[1], 6)
    return None


def lumber_quantity_m3(name: str, unit: str, qty: float | None) -> float | None:
    if qty is None:
        return None
    unit_key = normalized_unit(unit)
    if unit_key == "m3":
        return qty
    piece_volume = lumber_piece_volume_m3_from_text(name)
    cross_section = lumber_cross_section_m2_from_text(name)
    pack_count = package_count_from_text(name)
    pack_m = package_linear_m_from_text(name)
    if unit_key in {"pcs", "sheet"} and piece_volume:
        return round(qty * piece_volume, 6)
    if unit_key == "pack":
        if pack_count and piece_volume:
            return round(qty * pack_count * piece_volume, 6)
        if pack_m and cross_section:
            return round(qty * pack_m * cross_section, 6)
        if piece_volume:
            return round(qty * piece_volume, 6)
    if unit_key == "m" and cross_section:
        return round(qty * cross_section, 6)
    return None


@lru_cache(maxsize=40000)
def count_per_kg_from_text(text: str) -> float | None:
    source = clean_text(text).lower().replace(",", ".")
    patterns = [
        r"(\d+(?:\.\d+)?)\s*шт\s*/\s*кг",
        r"(\d+(?:\.\d+)?)\s*шт\s+в\s+кг",
        r"(\d+(?:\.\d+)?)\s*шт\s+на\s+кг",
    ]
    for pattern in patterns:
        match = re.search(pattern, source)
        if match:
            value = parse_number(match.group(1))
            if value:
                return value
    return None


@lru_cache(maxsize=40000)
def rebar_kg_per_piece_from_text(text: str) -> float | None:
    source = clean_text(text).lower().replace(",", ".")
    if "армат" not in source and "а500" not in source and "a500" not in source:
        return None
    diameter_match = re.search(r"(?:ф|d|f|ø)\s*(\d+(?:\.\d+)?)", source)
    if not diameter_match:
        diameter_match = re.search(r"(\d+(?:\.\d+)?)\s*мм", source)
    length_match = re.search(r"(?:l|длина)\s*=?\s*(\d+(?:\.\d+)?)\s*м", source)
    if not length_match:
        length_match = re.search(r"(\d+(?:\.\d+)?)\s*м\b", source)
    if not diameter_match:
        return None
    diameter = parse_number(diameter_match.group(1))
    length = parse_number(length_match.group(1)) if length_match else 11.7
    if not diameter or not length:
        return None
    return round(0.006165 * diameter * diameter * length, 6)


def first_number(*values: float | None) -> float | None:
    return next((value for value in values if value), None)


def converted_quantity(request: RequestItem, offer: SupplierItem) -> tuple[float | None, str]:
    if offer.qty is None:
        return None, ""
    req_unit = normalized_unit(request.unit)
    offer_unit = normalized_unit(offer.unit)

    offer_area = area_m2_from_text(offer.name)
    offer_dimension_area = dimension_area_m2_from_text(offer.name)
    request_area = area_m2_from_text(request.name)
    request_dimension_area = dimension_area_m2_from_text(request.name)
    offer_volume = volume_m3_from_text(offer.name)
    request_volume = volume_m3_from_text(request.name)
    offer_length = length_m_from_text(offer.name)
    request_length = length_m_from_text(request.name)
    offer_long_dimension = long_dimension_m_from_text(offer.name)
    request_long_dimension = long_dimension_m_from_text(request.name)
    offer_pack_count = package_count_from_text(offer.name)
    request_pack_count = package_count_from_text(request.name)
    offer_pack_kg = package_kg_from_text(offer.name)
    request_pack_kg = package_kg_from_text(request.name)
    offer_pack_m = package_linear_m_from_text(offer.name)
    request_pack_m = package_linear_m_from_text(request.name)
    combined_text = f"{offer.name} {request.name}"

    if is_lumber_text(combined_text):
        offer_m3 = lumber_quantity_m3(offer.name, offer.unit, offer.qty)
        request_m3 = lumber_quantity_m3(request.name, request.unit, request.qty)
        if offer_m3 is not None and (request_m3 is not None or req_unit == "m3"):
            return offer_m3, "m3"

    if req_unit == offer_unit:
        return offer.qty, req_unit

    if req_unit == "pcs" and offer_unit == "pack":
        count = first_number(offer_pack_count, request_pack_count)
        if count:
            return offer.qty * count, "pcs"
    if offer_unit == "pcs" and req_unit == "pack":
        count = first_number(request_pack_count, offer_pack_count)
        if count:
            return offer.qty / count, "pack"
    if req_unit == "kg" and offer_unit == "pack":
        kg = first_number(offer_pack_kg, request_pack_kg)
        if kg:
            return offer.qty * kg, "kg"
    if offer_unit == "kg" and req_unit == "pack":
        kg = first_number(request_pack_kg, offer_pack_kg)
        if kg:
            return offer.qty / kg, "pack"
    if req_unit == "kg" and offer_unit in {"pcs", "sheet"}:
        kg = first_number(offer_pack_kg, request_pack_kg)
        if kg:
            return offer.qty * kg, "kg"
    if offer_unit == "kg" and req_unit in {"pcs", "sheet"}:
        kg = first_number(request_pack_kg, offer_pack_kg)
        if kg:
            return offer.qty / kg, req_unit
    if req_unit == "m" and offer_unit == "pack":
        meters = first_number(offer_pack_m, request_pack_m)
        if meters:
            return offer.qty * meters, "m"
    if offer_unit == "m" and req_unit == "pack":
        meters = first_number(request_pack_m, offer_pack_m)
        if meters:
            return offer.qty / meters, "pack"
    if req_unit == "m" and offer_unit in {"pcs", "sheet"}:
        length = first_number(offer_length, request_length, offer_long_dimension, request_long_dimension)
        if length:
            return offer.qty * length, "m"
    if offer_unit == "m" and req_unit in {"pcs", "sheet"}:
        length = first_number(request_length, offer_length, request_long_dimension, offer_long_dimension)
        if length:
            return offer.qty / length, req_unit
    if req_unit == "m2" and offer_unit in {"pcs", "sheet"}:
        area = first_number(offer_dimension_area, offer_area, request_dimension_area, request_area)
        if area:
            return offer.qty * area, "m2"
    if req_unit == "m2" and offer_unit in {"roll", "pack"}:
        area = first_number(offer_area, offer_dimension_area, request_area, request_dimension_area)
        if area:
            return offer.qty * area, "m2"
    if offer_unit == "m2" and req_unit in {"pcs", "sheet"}:
        area = first_number(request_dimension_area, request_area, offer_dimension_area, offer_area)
        if area:
            return offer.qty / area, req_unit
    if offer_unit == "m2" and req_unit in {"roll", "pack"}:
        area = first_number(request_area, request_dimension_area, offer_area, offer_dimension_area)
        if area:
            return offer.qty / area, req_unit
    if req_unit == "m3" and offer_unit in {"pcs", "sheet", "pack"}:
        volume = first_number(offer_volume, request_volume)
        if volume:
            return offer.qty * volume, "m3"
    if offer_unit == "m3" and req_unit in {"pcs", "sheet", "pack"}:
        volume = first_number(request_volume, offer_volume)
        if volume:
            return offer.qty / volume, req_unit
    if req_unit == "pcs" and offer_unit == "kg":
        count_per_kg = first_number(count_per_kg_from_text(offer.name), count_per_kg_from_text(request.name))
        if count_per_kg:
            return offer.qty * count_per_kg, "pcs"
    if req_unit == "ton":
        if offer_unit == "kg":
            return offer.qty / 1000, "ton"
        if offer_unit == "pcs":
            kg_per_piece = first_number(
                rebar_kg_per_piece_from_text(offer.name),
                rebar_kg_per_piece_from_text(request.name),
                rebar_kg_per_piece_from_text(combined_text),
            )
            if kg_per_piece:
                return offer.qty * kg_per_piece / 1000, "ton"
    if req_unit == "kg" and offer_unit == "pcs":
        kg_per_piece = first_number(
            rebar_kg_per_piece_from_text(offer.name),
            rebar_kg_per_piece_from_text(request.name),
            rebar_kg_per_piece_from_text(combined_text),
        )
        if kg_per_piece:
            return offer.qty * kg_per_piece, "kg"
    return None, ""


def fmt_quantity_value(value: float | None) -> str:
    if value is None:
        return ""
    if abs(value - round(value)) < 0.001:
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def linear_meter_quantity_check(request: RequestItem, offer: SupplierItem) -> QuantityCheck | None:
    if request.qty is None or offer.qty is None:
        return None
    req_unit = normalized_unit(request.unit)
    offer_unit = normalized_unit(offer.unit)
    linear_units = {"m", "pcs", "sheet", "pack"}
    if req_unit not in linear_units or offer_unit not in linear_units or not ({req_unit, offer_unit} & {"pcs", "sheet"}):
        return None
    combined_text = f"{request.name} {offer.name}"
    if not is_linear_profile_text(combined_text):
        return None
    length = first_number(
        length_m_from_text(offer.name),
        length_m_from_text(request.name),
        long_dimension_m_from_text(offer.name),
        long_dimension_m_from_text(request.name),
    )
    request_pack_m = package_linear_m_from_text(request.name)
    offer_pack_m = package_linear_m_from_text(offer.name)
    if not length and not request_pack_m and not offer_pack_m:
        return None
    request_m = request.qty
    if req_unit in {"pcs", "sheet"} and length:
        request_m = request.qty * length
    elif req_unit == "pack" and request_pack_m:
        request_m = request.qty * request_pack_m
    offer_m = offer.qty
    if offer_unit in {"pcs", "sheet"} and length:
        offer_m = offer.qty * length
    elif offer_unit == "pack" and offer_pack_m:
        offer_m = offer.qty * offer_pack_m
    display = f"{fmt_qty(offer.qty)} {offer.unit}".strip()
    if offer_unit in {"pcs", "sheet", "pack"}:
        unit_label = request.unit if normalized_unit(request.unit) == "m" else "м.п."
        display = f"{display}\n~ {fmt_quantity_value(offer_m)} {unit_label}".strip()
    tolerance = max(0.01, request_m * 0.02)
    if abs(request_m - offer_m) <= tolerance:
        return QuantityCheck("ok", display, offer_m, "m")
    if offer_m < request_m:
        return QuantityCheck("low", display, offer_m, "m")
    return QuantityCheck("high", display, offer_m, "m")


def lumber_quantity_check(request: RequestItem, offer: SupplierItem) -> QuantityCheck | None:
    if request.qty is None or offer.qty is None:
        return None
    combined_text = f"{request.name} {offer.name}"
    if not is_lumber_text(combined_text):
        return None
    request_m3 = lumber_quantity_m3(request.name, request.unit, request.qty)
    offer_m3 = lumber_quantity_m3(offer.name, offer.unit, offer.qty)
    if request_m3 is None or offer_m3 is None:
        return None
    original = f"{fmt_qty(offer.qty)} {offer.unit}".strip()
    display = f"{original}\n~ {fmt_quantity_value(offer_m3)} м3".strip()
    tolerance = max(0.001, request_m3 * 0.02)
    if abs(request_m3 - offer_m3) <= tolerance:
        return QuantityCheck("ok", display, offer_m3, "m3")
    if offer_m3 < request_m3:
        return QuantityCheck("low", display, offer_m3, "m3")
    return QuantityCheck("high", display, offer_m3, "m3")


def quantity_check(request: RequestItem, offer: SupplierItem) -> QuantityCheck:
    original = f"{fmt_qty(offer.qty)} {offer.unit}".strip()
    if offer.override_qty is not None and offer.override_qty > 0 and offer.override_unit:
        display = f"{original}\n~ {fmt_quantity_value(offer.override_qty)} {offer.override_unit}".strip()
        req_unit = normalized_unit(request.unit)
        override_unit = normalized_unit(offer.override_unit)
        if request.qty is None or req_unit != override_unit:
            return QuantityCheck("unknown", display, offer.override_qty, override_unit)
        tolerance = max(0.01, request.qty * 0.02)
        if abs(request.qty - offer.override_qty) <= tolerance:
            return QuantityCheck("ok", display, offer.override_qty, override_unit)
        if offer.override_qty < request.qty:
            return QuantityCheck("low", display, offer.override_qty, override_unit)
        return QuantityCheck("high", display, offer.override_qty, override_unit)
    if request.qty is None or offer.qty is None:
        return QuantityCheck("unknown", original)
    lumber_check = lumber_quantity_check(request, offer)
    if lumber_check:
        return lumber_check
    linear_check = linear_meter_quantity_check(request, offer)
    if linear_check:
        return linear_check
    req_unit = normalized_unit(request.unit)
    offer_unit = normalized_unit(offer.unit)
    converted, converted_unit = converted_quantity(request, offer)
    display = original
    if converted is not None and converted_unit == req_unit and (req_unit != offer_unit or abs(converted - offer.qty) > 0.001):
        display = f"{original}\n~ {fmt_quantity_value(converted)} {request.unit}".strip()
    comparable_qty = converted if converted is not None and converted_unit == req_unit else (offer.qty if req_unit == offer_unit else None)
    if comparable_qty is None:
        return QuantityCheck("unknown", display)
    tolerance = max(0.01, request.qty * 0.02)
    if abs(request.qty - comparable_qty) <= tolerance:
        return QuantityCheck("ok", display, comparable_qty, req_unit)
    if comparable_qty < request.qty:
        return QuantityCheck("low", display, comparable_qty, req_unit)
    return QuantityCheck("high", display, comparable_qty, req_unit)


def quantities_compatible(request: RequestItem, offer: SupplierItem) -> bool:
    return quantity_check(request, offer).status == "ok"


@lru_cache(maxsize=40000)
def numeric_tokens(text: str) -> frozenset[str]:
    return frozenset(m.group(0).replace(",", ".") for m in re.finditer(r"\d+(?:[,.]\d+)?", text))


@lru_cache(maxsize=40000)
def brands(text: str) -> frozenset[str]:
    source = normalize(text)
    found = set()
    for raw, canonical in KNOWN_BRANDS.items():
        if raw in source:
            found.add(canonical)
    return frozenset(found)


def match_score(request_name: str, supplier_name: str) -> tuple[float, str]:
    if product_types_incompatible(request_name, supplier_name):
        return 0.0, "несовместимые товарные категории"

    req_norm = normalize(request_name)
    sup_norm = normalize(supplier_name)
    seq = SequenceMatcher(None, req_norm, sup_norm).ratio()

    req_tokens = tokens(request_name)
    sup_tokens = tokens(supplier_name)
    union = req_tokens | sup_tokens
    jaccard = len(req_tokens & sup_tokens) / len(union) if union else 0

    req_nums = numeric_tokens(request_name)
    sup_nums = numeric_tokens(supplier_name)
    num_union = req_nums | sup_nums
    num_score = len(req_nums & sup_nums) / len(num_union) if num_union else 0.4
    req_categories = item_categories(request_name)
    sup_categories = item_categories(supplier_name)
    category_overlap = bool(req_categories & sup_categories)

    req_brands = brands(request_name)
    sup_brands = brands(supplier_name)
    brand_conflict = bool(req_brands and sup_brands and req_brands.isdisjoint(sup_brands))

    score = 0.42 * seq + 0.43 * jaccard + 0.15 * num_score
    if category_overlap:
        score += 0.14
    if brand_conflict:
        score -= 0.22

    reasons = []
    if brand_conflict:
        reasons.append(f"разные бренды: {', '.join(sorted(req_brands))} / {', '.join(sorted(sup_brands))}")
    if score < 0.62:
        reasons.append("низкая уверенность сопоставления")
    return max(0, min(1, score)), "; ".join(reasons)


def find_header_row(ws, required_terms: Iterable[str]) -> tuple[int, dict[str, int]] | None:
    required = [term.lower() for term in required_terms]
    for row in ws.iter_rows():
        values = [clean_text(cell.value).lower() for cell in row]
        joined = " ".join(values)
        if all(term in joined for term in required):
            return row[0].row, {values[i]: i + 1 for i in range(len(values)) if values[i]}
    return None


def find_column(ws, header_row: int, variants: Iterable[str]) -> int | None:
    variants = [v.lower() for v in variants]
    for cell in ws[header_row]:
        value = clean_text(cell.value).lower()
        if any(variant in value for variant in variants):
            return cell.column
    return None


def find_column_preferred(
    ws,
    header_row: int,
    exact: Iterable[str],
    contains: Iterable[str],
    exclude: Iterable[str] = (),
) -> int | None:
    exact_values = [value.lower() for value in exact]
    contains_values = [value.lower() for value in contains]
    exclude_values = [value.lower() for value in exclude]
    fallback: int | None = None
    for cell in ws[header_row]:
        value = clean_text(cell.value).lower()
        if not value or any(marker in value for marker in exclude_values):
            continue
        if value in exact_values:
            return cell.column
        if fallback is None and any(marker in value for marker in contains_values):
            fallback = cell.column
    return fallback


def read_request_sheet(ws) -> list[RequestItem] | None:
    found = find_header_row(ws, ["описание", "количество"])
    if not found:
        found = find_header_row(ws, ["описание", "объем"])
    if not found:
        found = find_header_row(ws, ["позици", "кол-во"])
    if not found:
        found = find_header_row(ws, ["позици", "ед."])
    if not found:
        found = find_header_row(ws, ["наименование", "кол-во"])
    if not found:
        found = find_header_row(ws, ["наименование", "ед."])
    if not found:
        found = find_header_row(ws, ["наименование", "количество"])
    if not found:
        found = find_header_row(ws, ["наименование", "количесв"])
    if not found:
        found = find_header_row(ws, ["наименование", "единиц"])
    if not found:
        return None

    header_row, _ = found
    pos_col = find_column(ws, header_row, ["№", "номер"]) or 1
    name_col = find_column(ws, header_row, ["описание закупаемой", "описание", "позици", "наименование", "товар"])
    specs_col = find_column(ws, header_row, ["технические характеристики", "гост"])
    unit_col = find_column(ws, header_row, ["единица измерения", "единицы измерения", "ед. измерения", "ед измерения", "ед. из", "ед из", "ед."])
    qty_col = find_column(ws, header_row, ["необходимый объем", "количество", "количесвто", "количесв", "кол-во", "кол во", "кол-во."])

    if not name_col or not qty_col:
        return None

    items: list[RequestItem] = []
    for row in range(header_row + 1, ws.max_row + 1):
        pos = clean_text(ws.cell(row, pos_col).value) if pos_col != name_col else ""
        name = clean_text(ws.cell(row, name_col).value)
        unit = clean_text(ws.cell(row, unit_col).value) if unit_col else ""
        qty = parse_number(ws.cell(row, qty_col).value) if qty_col else None
        specs = clean_text(ws.cell(row, specs_col).value) if specs_col else ""
        if not name:
            continue
        if qty is None and not unit:
            continue
        if not pos:
            pos = str(len(items) + 1)
        items.append(RequestItem(pos=pos, name=name, specs=specs, unit=unit, qty=qty))
    return items or None


def read_request_xlsx(path: Path) -> list[RequestItem]:
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        items = read_request_sheet(ws)
        if items:
            return items
    raise ValueError(f"Не нашел строку заголовков заявки в {path.name}")


def supplier_name_from_file(path: Path) -> str:
    stem = path.stem
    known = {
        "петрович": "Петрович",
        "petrovich": "Петрович",
        "авангард": "Авангард-строй",
        "avangard": "Авангард-строй",
        "1350": "Авангард-строй",
        "еврострой": "Евростройгрупп",
        "eurostroy": "Евростройгрупп",
        "euro": "Евростройгрупп",
        "грандлайн": "Грандлайн",
        "grandline": "Грандлайн",
    }
    lower = stem.lower()
    for marker, name in known.items():
        if marker in lower:
            return name
    return stem[:45]


def invoice_no_from_text(text: str, fallback: str = "") -> str:
    source = clean_text(text)
    patterns = [
        r"\bкп[-\s№Nn#]*([A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9/_\-.]{1,30})",
        r"(?:счет|сч[её]т|кп|коммерческое предложение|заказ)\s*(?:№|N|N°|No|#)?\s*([A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9/_\-.]{1,30})",
        r"(?:№|N|N°|No|#)\s*([A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9/_\-.]{1,30})",
    ]
    for pattern in patterns:
        match = re.search(pattern, source, flags=re.IGNORECASE)
        if match:
            value = clean_text(match.group(1)).strip(".,;:")
            if value:
                return value
    file_match = re.search(r"\d{3,}", Path(fallback).stem if fallback else "")
    if file_match:
        return file_match.group(0)
    return fallback


def invoice_no_from_workbook(wb, path: Path) -> str:
    samples: list[str] = []
    for ws in wb.worksheets[:2]:
        max_row = min(ws.max_row, 18)
        max_col = min(ws.max_column, 8)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                value = clean_text(cell.value)
                if value:
                    samples.append(value)
    return invoice_no_from_text(" ".join(samples), path.name)


def invoice_label(offer: SupplierItem) -> str:
    return offer.invoice_no or invoice_no_from_text("", offer.source)


def supplier_invoice_summary(supplier: str, matches: list[Match]) -> str:
    labels: list[str] = []
    for match in matches:
        offer = match.supplier_item
        if offer.supplier != supplier:
            continue
        label = invoice_label(offer)
        if label and label not in labels:
            labels.append(label)
    if not labels:
        return ""
    if len(labels) > 3:
        return ", ".join(labels[:3]) + f" +{len(labels) - 3}"
    return ", ".join(labels)


def header_values(ws, header_row: int) -> dict[int, str]:
    return {cell.column: clean_text(cell.value).lower() for cell in ws[header_row]}


def first_header_col(headers: dict[int, str], includes: Iterable[str], excludes: Iterable[str] = ()) -> int | None:
    includes = [value.lower() for value in includes]
    excludes = [value.lower() for value in excludes]
    for col, value in headers.items():
        if not value or any(marker in value for marker in excludes):
            continue
        if any(marker in value for marker in includes):
            return col
    return None


def looks_like_vat_column(ws, col: int, header_row: int) -> bool:
    values = []
    for row in range(header_row + 1, min(ws.max_row, header_row + 25) + 1):
        value = parse_number(ws.cell(row, col).value)
        if value is not None:
            values.append(value)
    if len(values) < 3:
        return False
    common_vat = sum(1 for value in values if value in {0, 10, 18, 20, 22})
    return common_vat / len(values) >= 0.7


def find_loose_supplier_header(ws) -> int | None:
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [clean_text(cell.value).lower() for cell in ws[row]]
        joined = " ".join(values)
        has_name = any(marker in joined for marker in ["товар", "наименование", "номенклатура"])
        has_price = "цена" in joined or "стоимость" in joined or "сумма" in joined
        has_qty = "количество" in joined or "кол-во" in joined or "кол во" in joined
        if has_name and has_price and has_qty:
            return row
    return None


def read_loose_supplier_xlsx(ws, path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    header_row = find_loose_supplier_header(ws)
    if not header_row:
        return []

    headers = header_values(ws, header_row)
    name_col = first_header_col(headers, ["номенклатура", "наименование", "товар"], ["код", "артикул"])
    qty_col = first_header_col(headers, ["количество", "кол-во", "кол во"])
    unit_col = first_header_col(headers, ["ед.", "единиц", "ед "])
    price_col = first_header_col(headers, ["цена"], ["ндс"])
    total_col = first_header_col(headers, ["стоимость", "сумма"], ["ндс"])
    pos_col = first_header_col(headers, ["№", "номер"]) or 1

    if not name_col:
        return []

    # CSV converted through Excel sometimes shifts data one column left when a
    # text header is split into "Товары (работы" / "услуги)".
    if price_col and looks_like_vat_column(ws, price_col, header_row) and price_col > 1:
        price_col -= 1
    if qty_col == price_col and qty_col and qty_col > 1:
        qty_col -= 1
    if not qty_col and price_col and price_col > name_col + 1:
        qty_col = price_col - 1
    if not price_col:
        price_col = first_header_col(headers, ["стоимость", "сумма"], ["ндс"])
    if not qty_col or not price_col:
        return []

    items: list[SupplierItem] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = clean_text(ws.cell(row, name_col).value)
        if not name or "итого" in name.lower() or "всего" in name.lower():
            continue
        price = parse_number(ws.cell(row, price_col).value)
        total = parse_number(ws.cell(row, total_col).value) if total_col else None
        qty = parse_quantity(ws.cell(row, qty_col).value, price, total)
        if price is None and total is None:
            continue
        if qty is None and price is None:
            continue
        items.append(
            SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=clean_text(ws.cell(row, pos_col).value) or str(row),
                name=name,
                qty=qty,
                unit=clean_text(ws.cell(row, unit_col).value) if unit_col else "",
                price=price,
                total=total,
                invoice_no=invoice_no,
            )
        )
    return items


def read_supplier_xlsx(path: Path, supplier: str | None = None) -> list[SupplierItem]:
    wb = load_workbook(path, data_only=True)
    supplier = supplier or supplier_name_from_file(path)
    invoice_no = invoice_no_from_workbook(wb, path)
    items: list[SupplierItem] = []

    for ws in wb.worksheets:
        table_header = find_header_row(ws, ["товар", "кол-во"])
        if table_header:
            header_row, _ = table_header
            name_col = find_column_preferred(
                ws,
                header_row,
                exact=["товар", "наименование"],
                contains=["номенклатура", "наименование", "товар"],
                exclude=["код товара", "код"],
            )
            qty_col = find_column_preferred(ws, header_row, exact=["кол-во", "количество"], contains=["кол-во", "количество"])
            unit_col = find_column_preferred(ws, header_row, exact=["ед.", "ед"], contains=["ед."])
            price_col = find_column_preferred(ws, header_row, exact=["цена"], contains=["цена"])
            total_col = find_column(ws, header_row, ["сумма", "стоимость"])
            pos_col = find_column(ws, header_row, ["№"]) or 1
            if name_col and qty_col and price_col:
                for row in range(header_row + 1, ws.max_row + 1):
                    name = clean_text(ws.cell(row, name_col).value)
                    if not name or "итого" in name.lower() or "всего" in name.lower():
                        continue
                    price = parse_number(ws.cell(row, price_col).value)
                    total = parse_number(ws.cell(row, total_col).value) if total_col else None
                    if price is None and total is None:
                        continue
                    items.append(
                        SupplierItem(
                            supplier=supplier,
                            source=path.name,
                            row_no=clean_text(ws.cell(row, pos_col).value) or str(row),
                            name=name,
                            qty=parse_quantity(ws.cell(row, qty_col).value, price, total),
                            unit=clean_text(ws.cell(row, unit_col).value) if unit_col else "",
                            price=price,
                            total=total,
                            invoice_no=invoice_no,
                        )
                    )
                continue

        request_like_header = find_header_row(ws, ["описание", "предельная цена"])
        if request_like_header:
            header_row, _ = request_like_header
            pos_col = find_column(ws, header_row, ["№"]) or 1
            name_col = find_column(ws, header_row, ["описание закупаемой", "описание"])
            qty_col = find_column(ws, header_row, ["необходимый объем", "количество"])
            unit_col = find_column(ws, header_row, ["ед. измерения", "ед измерения"])
            price_col = find_column(ws, header_row, ["предельная цена", "цена"])
            total_col = find_column(ws, header_row, ["стоимость", "сумма"])
            delivery_col = find_column(ws, header_row, ["срок"])
            if name_col and price_col:
                for row in range(header_row + 1, ws.max_row + 1):
                    name = clean_text(ws.cell(row, name_col).value)
                    price = parse_number(ws.cell(row, price_col).value)
                    total = parse_number(ws.cell(row, total_col).value) if total_col else None
                    if not name or (price is None and total is None):
                        continue
                    items.append(
                        SupplierItem(
                            supplier=supplier,
                            source=path.name,
                            row_no=clean_text(ws.cell(row, pos_col).value) or str(row),
                            name=name,
                            qty=parse_quantity(ws.cell(row, qty_col).value, price, total) if qty_col else None,
                            unit=clean_text(ws.cell(row, unit_col).value) if unit_col else "",
                            price=price,
                            total=total,
                            delivery=clean_text(ws.cell(row, delivery_col).value) if delivery_col else "",
                            invoice_no=invoice_no,
                        )
                    )
        if not items:
            items.extend(read_loose_supplier_xlsx(ws, path, supplier, invoice_no))
    return items


def read_supplier_xls(path: Path, supplier: str | None = None) -> list[SupplierItem]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise ValueError(
            f"{path.name}: для чтения старого .xls не установлена библиотека xlrd. "
            "Сохраните файл как .xlsx или PDF и запустите повторно."
        ) from exc

    book = xlrd.open_workbook(str(path))
    supplier = supplier or supplier_name_from_file(path)
    items: list[SupplierItem] = []

    def cell_text(sheet, row: int, col: int | None) -> str:
        if col is None or col < 0:
            return ""
        value = sheet.cell_value(row, col)
        if sheet.cell_type(row, col) == xlrd.XL_CELL_DATE:
            try:
                parts = xlrd.xldate_as_tuple(value, book.datemode)
                return f"{parts[2]:02d}/{parts[1]:02d}/{str(parts[0])[-2:]}"
            except Exception:
                return clean_text(value)
        return clean_text(value)

    def first_col(headers: dict[int, str], includes: Iterable[str], excludes: Iterable[str] = ()) -> int | None:
        includes = [value.lower() for value in includes]
        excludes = [value.lower() for value in excludes]
        for col, value in headers.items():
            if not value or any(marker in value for marker in excludes):
                continue
            if any(marker in value for marker in includes):
                return col
        return None

    for sheet in book.sheets():
        header_row: int | None = None
        for row in range(min(sheet.nrows, 80)):
            values = [clean_text(sheet.cell_value(row, col)).lower() for col in range(sheet.ncols)]
            joined = " ".join(values)
            has_name = any(marker in joined for marker in ["имя товара", "наименование", "товар", "номенклатура"])
            has_qty = any(marker in joined for marker in ["кол-во", "количество", "кол во"])
            has_price = any(marker in joined for marker in ["цена", "сумма", "стоимость"])
            if has_name and has_qty and has_price:
                header_row = row
                break
        if header_row is None:
            continue

        headers = {col: clean_text(sheet.cell_value(header_row, col)).lower() for col in range(sheet.ncols)}
        name_col = first_col(headers, ["имя товара", "наименование", "номенклатура", "товар"], ["код", "артикул"])
        qty_col = first_col(headers, ["кол-во", "количество", "кол во"])
        unit_col = first_col(headers, ["ед.изм", "ед. изм", "единиц", "ед."])
        price_col = first_col(headers, ["цена"], ["ндс %"])
        total_col = first_col(headers, ["сумма", "стоимость"], ["ндс %"])
        delivery_col = first_col(headers, ["срок поставки", "срок"])
        pos_col = first_col(headers, ["№", "n", "номер"])

        if name_col is None:
            continue

        for row in range(header_row + 1, sheet.nrows):
            name = cell_text(sheet, row, name_col)
            if not name:
                continue
            lower = name.lower()
            if "итого" in lower or "всего" in lower:
                continue
            qty = parse_number(sheet.cell_value(row, qty_col)) if qty_col is not None else None
            price = parse_number(sheet.cell_value(row, price_col)) if price_col is not None else None
            total = parse_number(sheet.cell_value(row, total_col)) if total_col is not None else None
            if price is None and total is None:
                continue
            if price is None and total is not None and qty:
                price = round(total / qty, 4)
            if total is None and price is not None and qty:
                total = round(price * qty, 4)
            items.append(
                SupplierItem(
                    supplier=supplier,
                    source=path.name,
                    row_no=cell_text(sheet, row, pos_col) if pos_col is not None else str(row - header_row),
                    name=name,
                    qty=qty,
                    unit=cell_text(sheet, row, unit_col),
                    price=price,
                    total=total,
                    delivery=cell_text(sheet, row, delivery_col),
                    invoice_no=path.stem,
                )
            )

    return items


ROW_RE = re.compile(
    r"^\s*(?P<row>\d+)\s+(?:(?P<code>\d{4,}|[0-9]{4,}\s+заказ)\s+)?"
    r"(?P<name>.*?)\s+(?P<qty>\d[\d\s]*(?:[,.]\d+)?)\s+"
    r"(?P<unit>[A-Za-zА-Яа-яЁё.]+)\s+"
    r"(?P<money>\d[\d\s]*,\d{2}(?:\s+\d[\d\s]*,\d{2}){1,3})\s*$"
)


def clean_pdf_name_line(line: str) -> str:
    line = clean_text(line)
    line = re.sub(r"^\d{5,}\s+", "", line)
    line = re.sub(r"^\d{5,}\s+заказ\s+", "", line)
    return line


def is_product_continuation(line: str) -> bool:
    text = clean_text(line)
    if not text:
        return False
    lower = text.lower()
    if any(marker in lower for marker in ["итого", "ндс", "всего к оплате", "страница", "поставщик", "покупатель"]):
        return False
    if re.search(r"\d[\d\s]*,\d{2}\s+\d[\d\s]*,\d{2}", text):
        return False
    return bool(re.search(r"[A-Za-zА-Яа-яЁё]", text))


def pdf_text_lines(path: Path, layout: bool = True) -> list[str]:
    reader = PdfReader(str(path))
    text_parts = []
    for page in reader.pages:
        try:
            text_parts.append(page.extract_text(extraction_mode="layout") or "")
        except TypeError:
            text_parts.append(page.extract_text() or "")
    return "\n".join(text_parts).splitlines()


def delivery_from_pdf_lines(lines: list[str]) -> str:
    for line in lines:
        clean = clean_text(line)
        lower = clean.lower()
        if "срок исполнения" in lower or "срок поставки" in lower or "ориентировочный срок поставки" in lower:
            parts = re.split(r":", clean, maxsplit=1)
            return clean_text(parts[-1]) if len(parts) > 1 else clean
    return ""


def parse_amount_token(value: str) -> float | None:
    value = clean_text(value).replace("'", "").replace("`", "")
    value = re.sub(r"(?<=\d)\s+(?=\d)", "", value)
    return parse_number(value)


def money_values_from_text(text: str) -> list[float]:
    values: list[float] = []
    amount_re = re.compile(
        r"(?<![\w])("
        r"\d{1,3}(?:[ '\u202f`]\d{3})+(?:[,.]\d{2})?"
        r"|\d{4}\s+\d{3}(?:\s+\d{3})*(?:[,.]\d{2})?"
        r"|\d{5,}(?:[,.]\d{2})"
        r")(?![\w])"
    )
    for match in amount_re.finditer(clean_text(text)):
        value = parse_amount_token(match.group(0))
        if value is not None and 10_000 <= value <= 100_000_000:
            values.append(value)
    return values


def first_amount_from_text(text: str) -> float | None:
    match = re.search(r"\d[\d\s]*(?:[,.]\d{2})", clean_text(text))
    if not match:
        return None
    return parse_amount_token(match.group(0))


def looks_like_unit_token(text: str) -> bool:
    lower = clean_text(text).lower()
    return any(unit in lower for unit in ["шт", "м", "кг", "уп", "комп", "усл"])


def is_offer_note_line(text: str) -> bool:
    lower = clean_text(text).lower()
    if not lower:
        return True
    note_markers = [
        "в наличии",
        "под заказ",
        "кратно",
        "адрес доставки",
        "итого",
        "ндс",
        "счет-договор",
        "продолжение",
        "страница",
        "наименование",
        "кол-во",
        "срок",
        "цена",
        "сумма",
        "(дни)",
        "(руб)",
    ]
    return any(marker in lower for marker in note_markers)


def parse_split_pdf_table(lines: list[str], path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    items: list[SupplierItem] = []
    current: SupplierItem | None = None
    delivery = delivery_from_pdf_lines(lines)

    def flush_current() -> None:
        nonlocal current
        if current and current.name and "доставка" not in current.name.lower():
            items.append(current)
        current = None

    for raw_line in lines:
        parts = [part for part in re.split(r"\s{2,}", raw_line.strip()) if part]
        if not parts:
            continue

        if parts[0].isdigit() and len(parts) >= 6:
            row_no = parts[0]
            name = clean_pdf_name_line(parts[1])
            if name and is_service_text(name):
                amounts = [first_amount_from_text(part) for part in parts[2:]]
                amounts = [amount for amount in amounts if amount is not None]
                if amounts:
                    flush_current()
                    total = amounts[-1]
                    items.append(
                        SupplierItem(
                            supplier=supplier,
                            source=path.name,
                            row_no=row_no,
                            name=name,
                            qty=1,
                            unit="усл",
                            price=total,
                            total=total,
                            delivery=delivery,
                            invoice_no=invoice_no,
                        )
                    )
                continue
            if not name or "доставка" in name.lower():
                flush_current()
                continue

            qty: float | None = None
            unit = ""
            price: float | None = None
            total: float | None = None

            if len(parts) >= 9 and looks_like_unit_token(parts[3]) and first_amount_from_text(parts[5]) is not None:
                unit = clean_text(parts[3]).rstrip(".")
                price = first_amount_from_text(parts[5])
                total = first_amount_from_text(parts[8]) or first_amount_from_text(parts[6])
                qty = parse_quantity(parts[4], price, total)
            elif looks_like_unit_token(parts[3]) and first_amount_from_text(parts[-2]) is not None and first_amount_from_text(parts[-1]) is not None:
                unit = clean_text(parts[3]).rstrip(".")
                price = first_amount_from_text(parts[-2])
                total = first_amount_from_text(parts[-1])
                qty = parse_quantity(parts[2], price, total)

            if price is None and total is None:
                continue

            flush_current()
            current = SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=row_no,
                name=name,
                qty=qty,
                unit=unit,
                price=price,
                total=total,
                delivery=delivery,
                invoice_no=invoice_no,
            )
            continue

        if current and not is_offer_note_line(raw_line):
            continuation = clean_pdf_name_line(raw_line)
            if continuation and not re.match(r"^\d+\s+", continuation):
                current.name = clean_text(f"{current.name} {continuation}")

    flush_current()
    return items


OCR_PRODUCT_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("Блок ввода и узла учета тепловой энергии", re.compile(r"блок\s+ввода", re.IGNORECASE)),
    ("Блок системы ГВС", re.compile(r"блок\s+системы\s+гвс", re.IGNORECASE)),
    ("Блок системы отопления", re.compile(r"блок\s+системы\s+отоплен", re.IGNORECASE)),
    ("Распределительная гребенка", re.compile(r"распределительн\w*\s+греб", re.IGNORECASE)),
    ("Узел ввода", re.compile(r"\b(?:узел|ш)\s+ввода\b", re.IGNORECASE)),
    ("Тепловой пункт системы ГВС", re.compile(r"(?:тепловой\s+)?пункт\s+системы\s+гвс", re.IGNORECASE)),
    ("Тепловой пункт системы отопления", re.compile(r"(?:тепловой\s+)?пункт\s+системы\s+отоплен|тепловой\s+пункт.*?отоплен", re.IGNORECASE)),
    ("Тепловой пункт линии подпитки", re.compile(r"(?:тепловой\s+)?пункт\s+линии\s+подпитки|линии\s+подпитки", re.IGNORECASE)),
    ("Модульный коллектор", re.compile(r"модульн\w*|коллектор", re.IGNORECASE)),
    ("Шкаф автоматизации", re.compile(r"шкаф\s+автоматизации|шкаф\b|автоматизации", re.IGNORECASE)),
]


def ocr_product_label(line: str) -> str | None:
    clean = clean_text(line)
    for label, pattern in OCR_PRODUCT_PATTERNS:
        if pattern.search(clean):
            return label
    return None


def has_ocr_product_marker(line: str) -> bool:
    return ocr_product_label(line) is not None


def ocr_label_near_amount(lines: list[str], idx: int) -> str | None:
    for end in range(idx + 1, min(len(lines), idx + 4) + 1):
        label = ocr_product_label(clean_text(" ".join(lines[idx:end])))
        if label:
            return label
    for start in range(idx - 1, max(-1, idx - 3), -1):
        label = ocr_product_label(clean_text(" ".join(lines[start : idx + 1])))
        if label:
            return label
    return None


def parse_structured_ocr_pdf(lines: list[str], path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    items: list[SupplierItem] = []
    delivery = delivery_from_pdf_lines(lines)
    clean_lines = [clean_text(line) for line in lines]
    seen: set[tuple[str, int]] = set()

    for idx, line in enumerate(clean_lines):
        lower = line.lower()
        if any(word in lower for word in ["итого", "всего", "ндс", "доставка", "реквизит", "кои"]):
            continue
        values = money_values_from_text(line)
        if not values:
            continue
        label = ocr_label_near_amount(clean_lines, idx)
        if not label:
            continue
        total = values[-1]
        key = (label.lower(), round(total))
        if key in seen:
            continue
        seen.add(key)
        items.append(
            SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=str(len(items) + 1),
                name=label,
                qty=1,
                unit="шт",
                price=total,
                total=total,
                delivery=delivery,
                invoice_no=invoice_no,
            )
        )
    return items


def looks_like_product_text(text: str) -> bool:
    lower = clean_text(text).lower()
    markers = [
        "блок",
        "тепловой пункт",
        "узел",
        "гвс",
        "отоплен",
        "подпит",
        "коллектор",
        "шкаф",
        "битп",
        "ридан",
        "ув-",
        "бтп",
        "ангар",
    ]
    return any(marker in lower for marker in markers)


def clean_ocr_offer_name(text: str) -> str:
    clean = clean_text(text)
    clean = re.sub(r"\d(?:[\d\s'`]{2,})(?:[,.]\d{2})?", " ", clean)
    clean = re.sub(r"[$ВB]?S?0?203260034[-\w/]*", " ", clean, flags=re.IGNORECASE)
    clean = re.sub(r"\b\d+\s*\|?", " ", clean)
    clean = re.sub(r"\b(?:шт|руб|ндс|под заказ|вес|кг|недель)\b[:.]?", " ", clean, flags=re.IGNORECASE)
    clean = re.sub(r"[|\\[\\]{}]+", " ", clean)
    clean = re.sub(r"\s+", " ", clean).strip(" .,:;-")
    return clean


def source_row_no_from_table_prefix(text: str) -> str | None:
    clean = clean_text(text)
    match = re.match(r"^\s*(?P<row>\d{1,4})(?:\s+|$)", clean)
    if not match:
        return None
    row_no = int(match.group("row"))
    if row_no <= 0 or row_no > 9999:
        return None
    return str(row_no)


def parse_fuzzy_ocr_pdf(lines: list[str], path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    structured_items = parse_structured_ocr_pdf(lines, path, supplier, invoice_no)
    if structured_items:
        return structured_items

    items: list[SupplierItem] = []
    delivery = delivery_from_pdf_lines(lines)
    skip_words = ["итого", "всего", "ндс", "доставка", "стоимость доставки", "сумма", "реквизит"]
    clean_lines = [clean_text(line) for line in lines]

    for idx, line in enumerate(clean_lines):
        lower = line.lower()
        if any(word in lower for word in skip_words):
            continue
        values = money_values_from_text(line)
        if not values:
            continue

        context_parts = []
        for near_idx in range(max(0, idx - 2), min(len(clean_lines), idx + 3)):
            near = clean_lines[near_idx]
            near_lower = near.lower()
            if any(word in near_lower for word in ["итого", "всего", "ндс", "реквизит"]):
                continue
            context_parts.append(near)
        context = clean_text(" ".join(context_parts))
        if not looks_like_product_text(context):
            continue

        name = clean_ocr_offer_name(context)
        if len(name) < 8:
            continue
        total = values[-1]
        price = values[0]
        if total < 1000 or any(abs((offer.total or 0) - total) < 0.01 for offer in items):
            continue
        items.append(
            SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=str(len(items) + 1),
                name=name,
                qty=1,
                unit="шт",
                price=price,
                total=total,
                delivery=delivery,
                invoice_no=invoice_no,
            )
        )
    return items


def parse_generic_ocr_table_pdf(lines: list[str], path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    joined = clean_text(" ".join(lines)).lower()
    if "смит" not in joined and "товар" not in joined:
        return []

    product_marker = re.compile(
        r"(?:\bВК\b|\bПП\b|Кран|Муфта|Тройник|Труба|Уголок|Отвод|Заглушка|Клапан|Крестовина|"
        r"Крепление|Болт|Винт|Гайка|Герметик|Грунтовка|Подводка)",
        flags=re.IGNORECASE,
    )
    qty_pattern = re.compile(r"\b(?P<qty>\d{1,5})\s*(?:шт|шр|шо|ш\b|шт\.)", flags=re.IGNORECASE)
    decimal_amount = re.compile(r"(?<!\d)\d{1,7}(?:[,.]\d{2})(?!\d)")

    items: list[SupplierItem] = []
    seen: set[tuple[str, float | None]] = set()
    delivery = delivery_from_pdf_lines(lines)
    for raw_line in lines:
        line = clean_text(raw_line)
        if not product_marker.search(line):
            continue
        compact = re.sub(r"[\[\]|_]+", " ", line)
        compact = re.sub(r"\s+", " ", compact).strip()
        qty_match = qty_pattern.search(compact)
        if not qty_match:
            continue
        marker_match = product_marker.search(compact[: qty_match.start()])
        if not marker_match:
            continue

        source_row_no = source_row_no_from_table_prefix(compact[: marker_match.start()])
        name = clean_text(compact[marker_match.start() : qty_match.start()])
        name = re.sub(r"^\d{1,3}\s*\d{3,6}\s+", "", name)
        name = re.sub(r"^\d{3,6}\s+", "", name)
        name = name.strip(" -—Г")
        if len(name) < 6:
            continue

        qty = parse_number(qty_match.group("qty"))
        tail = compact[qty_match.end() :]
        amounts = [parse_amount_token(value) for value in decimal_amount.findall(tail)]
        amounts = [value for value in amounts if value is not None and value > 0]
        if not amounts:
            continue

        total = amounts[-1]
        price = amounts[0]
        if qty and total and (not price or price > total or abs(price * qty - total) > max(total * 0.35, 100)):
            price = round(total / qty, 4)

        key = (name.lower(), round(total, 2) if total is not None else None)
        if key in seen:
            continue
        seen.add(key)
        items.append(
            SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=source_row_no or str(len(items) + 1),
                name=name,
                qty=qty,
                unit="шт",
                price=price,
                total=total,
                delivery=delivery,
                invoice_no=invoice_no,
            )
        )
    return items


def parse_sib_k_pdf(lines: list[str], path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    items: list[SupplierItem] = []
    delivery = delivery_from_pdf_lines(lines)
    for idx, line in enumerate(lines, start=1):
        clean = clean_text(line)
        match = re.search(
            r"(?P<name>[A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\s]+?)\s*[–—-]\s*(?P<amount>\d[\d\s]*(?:[,.]\d{2})?)\s*р",
            clean,
            flags=re.IGNORECASE,
        )
        if not match:
            continue
        name = clean_text(match.group("name")).strip(" ,.;:")
        total = parse_amount_token(match.group("amount"))
        if not name or total is None:
            continue
        items.append(
            SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=str(len(items) + 1),
                name=name,
                qty=1,
                unit="шт",
                price=total,
                total=total,
                delivery=delivery,
                invoice_no=invoice_no,
            )
        )

    if items:
        return items

    joined = clean_text(" ".join(lines))
    total_match = re.search(r"за\s+(?P<amount>\d[\d\s]*(?:[,.]\d{2})?)\s*руб", joined, flags=re.IGNORECASE)
    if total_match and "битп" in joined.lower():
        total = parse_amount_token(total_match.group("amount"))
        if total is not None:
            return [
                SupplierItem(
                    supplier=supplier,
                    source=path.name,
                    row_no="1",
                    name="БИТП в составе: блок отопления, блок ГВС, блок УУТЭ, блок распределительных гребенок, изоляционные материалы, материалы для соединения между блоками",
                    qty=1,
                    unit="компл",
                    price=total,
                    total=total,
                    delivery=delivery,
                    invoice_no=invoice_no,
                )
            ]
    return []


def is_ridan_continuation(line: str) -> bool:
    clean = re.sub(r"\bкг\.?\b", " ", clean_text(line), flags=re.IGNORECASE)
    if not clean:
        return False
    lower = clean.lower()
    skip_markers = [
        "всего",
        "ндс",
        "условия",
        "срок:",
        "под заказ",
        "контакты",
        "гарантия",
        "цена указана",
        "доставка",
        "не для продажи",
        "настоящего коммерческого",
        "официальных партнеров",
        "расчет выполнил",
        "ответственный за объект",
        "обратившись по тел",
    ]
    if any(marker in lower for marker in skip_markers):
        return False
    return bool(re.search(r"[A-Za-zА-Яа-яЁё]", clean))


def clean_ridan_part(line: str) -> str:
    clean = re.sub(r"\bкг\.?\b", " ", clean_text(line), flags=re.IGNORECASE)
    clean = re.sub(r"\s+", " ", clean).strip(" ,.;:")
    return clean


def parse_ridan_pdf(lines: list[str], path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    items: list[SupplierItem] = []
    delivery = delivery_from_pdf_lines(lines)
    current: SupplierItem | None = None
    row_pattern = re.compile(
        r"^\s*(?P<row>\d+)\s+(?P<name>.+?)\s+"
        r"(?P<article>[A-ZА-Я]{1,3}\d[\w/.\-]*)\s+"
        r"(?P<price>\d[\d\s]*(?:[,.]\d{2})?)\s+"
        r"(?P<qty>\d+(?:[,.]\d+)?)\s+"
        r"(?P<total>\d[\d\s]*(?:[,.]\d{2})?)\b",
        flags=re.IGNORECASE,
    )

    for line in lines:
        clean = clean_text(line)
        lower = clean.lower()
        if "всего за товары" in lower or "для коммерческого предложения" in lower:
            break
        match = row_pattern.match(clean)
        if match:
            if current:
                items.append(current)
            price = parse_amount_token(match.group("price"))
            total = parse_amount_token(match.group("total"))
            row_delivery = delivery
            delivery_match = re.search(r"Срок:\s*([^:]+?)(?:\s+Вес|\s*$)", clean, flags=re.IGNORECASE)
            if delivery_match:
                row_delivery = clean_text(delivery_match.group(1))
            current = SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=match.group("row"),
                name=clean_ridan_part(match.group("name")),
                qty=parse_quantity(match.group("qty"), price, total),
                unit="шт",
                price=price,
                total=total,
                delivery=row_delivery,
                invoice_no=invoice_no,
            )
            continue
        if current and is_ridan_continuation(clean):
            part = clean_ridan_part(clean)
            if part:
                current.name = clean_text(current.name + " " + part)

    if current:
        items.append(current)
    return items


def parse_known_text_pdf(lines: list[str], path: Path, supplier: str, invoice_no: str) -> list[SupplierItem]:
    joined = clean_text(" ".join(lines)).lower()
    if "ридан" in joined or "тепловой пункт" in joined:
        items = parse_ridan_pdf(lines, path, supplier, invoice_no)
        if items:
            return items
    if "сиб-к" in joined or "битп" in joined:
        items = parse_sib_k_pdf(lines, path, supplier, invoice_no)
        if items:
            return items
    items = parse_generic_ocr_table_pdf(lines, path, supplier, invoice_no)
    if items:
        return items
    return parse_fuzzy_ocr_pdf(lines, path, supplier, invoice_no)


def configure_tesseract() -> str | None:
    found = shutil.which("tesseract")
    if found:
        return found
    common_paths = [
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]
    for candidate in common_paths:
        if candidate.exists():
            return str(candidate)
    return None


def local_tessdata_dir() -> Path:
    return Path(__file__).resolve().parent / ".tessdata"


def tesseract_safe_path(path: Path) -> str:
    resolved = str(path.resolve())
    if os.name != "nt":
        return resolved
    buffer = ctypes.create_unicode_buffer(260)
    result = ctypes.windll.kernel32.GetShortPathNameW(resolved, buffer, len(buffer))
    if result:
        return buffer.value
    return resolved


def available_tesseract_languages(tesseract_path: str) -> set[str]:
    try:
        import subprocess

        tessdata_dir = local_tessdata_dir()
        command = [tesseract_path]
        if tessdata_dir.exists():
            command.extend(["--tessdata-dir", tesseract_safe_path(tessdata_dir)])
        command.append("--list-langs")
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
    except Exception:
        return set()
    return {
        clean_text(line).strip()
        for line in result.stdout.splitlines()
        if clean_text(line).strip() and not line.lower().startswith("list of available")
    }


def ocr_pdf_lines(path: Path) -> list[str]:
    try:
        try:
            import pymupdf as fitz  # type: ignore
        except ModuleNotFoundError:
            import fitz  # type: ignore
        import pytesseract  # type: ignore
        from PIL import Image
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            f"{path.name}: PDF похож на скан/картинку. Для OCR нужны Python-пакеты PyMuPDF и pytesseract."
        ) from exc
    if not hasattr(fitz, "open"):
        raise RuntimeError(
            f"{path.name}: OCR не смог открыть PDF, потому что установлен конфликтующий модуль fitz. "
            "Переустановите PyMuPDF или запустите сервис через run_8001.bat."
        )

    tesseract_path = configure_tesseract()
    if not tesseract_path:
        raise RuntimeError(
            f"{path.name}: PDF похож на скан/картинку. Для OCR установите Tesseract OCR с русским языком."
        )
    if hasattr(pytesseract, "pytesseract"):
        pytesseract.pytesseract.tesseract_cmd = tesseract_path
    else:
        pytesseract.tesseract_cmd = tesseract_path
    languages = available_tesseract_languages(tesseract_path)
    if "rus" in languages and "eng" in languages:
        lang = "rus+eng"
    elif "rus" in languages:
        lang = "rus"
    else:
        lang = "eng"
    config = "--psm 6"
    tessdata_dir = local_tessdata_dir()
    if tessdata_dir.exists():
        config += f" --tessdata-dir {tesseract_safe_path(tessdata_dir)}"

    lines: list[str] = []
    with fitz.open(path) as doc:
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image = Image.open(io.BytesIO(pix.tobytes("png")))
            text = pytesseract.image_to_string(image, lang=lang, config=config)
            lines.extend(text.splitlines())
    return lines


def parse_layout_pdf(path: Path, supplier: str | None = None) -> list[SupplierItem]:
    supplier = supplier or supplier_name_from_file(path)
    lines = pdf_text_lines(path)
    if not any(clean_text(line) for line in lines):
        lines = ocr_pdf_lines(path)
    invoice_no = invoice_no_from_text(" ".join(lines[:35]), path.name)

    split_items = parse_split_pdf_table(lines, path, supplier, invoice_no)
    if len(split_items) >= 3:
        return split_items

    delivery = delivery_from_pdf_lines(lines)

    items: list[SupplierItem] = []
    in_table = False
    pending: list[str] = []
    current: SupplierItem | None = None

    def flush_pending_to_current() -> None:
        nonlocal pending, current
        if current and pending:
            current.name = clean_text(current.name + " " + " ".join(clean_pdf_name_line(x) for x in pending))
        pending = []

    for line in lines:
        lower = line.lower()
        if "товары" in lower and "сумма" in lower:
            in_table = True
            pending = []
            continue
        if not in_table:
            continue
        if "итого:" in lower or "всего к оплате" in lower:
            break

        match = ROW_RE.match(line)
        if match:
            if current:
                items.append(current)
            money_values = [money_to_float(x) for x in re.findall(r"\d[\d\s]*,\d{2}", match.group("money"))]
            money_values = [x for x in money_values if x is not None]
            price = money_values[1] if len(money_values) >= 3 else (money_values[0] if money_values else None)
            total = money_values[-1] if money_values else None
            name_parts = [clean_pdf_name_line(x) for x in pending]
            name_parts.append(clean_pdf_name_line(match.group("name")))
            current = SupplierItem(
                supplier=supplier,
                source=path.name,
                row_no=match.group("row"),
                name=clean_text(" ".join(name_parts)),
                qty=parse_quantity(match.group("qty"), price, total),
                unit=clean_text(match.group("unit")),
                price=price,
                total=total,
                delivery=delivery,
                invoice_no=invoice_no,
            )
            pending = []
            continue

        if current and is_product_continuation(line):
            current.name = clean_text(current.name + " " + clean_pdf_name_line(line))
        elif is_product_continuation(line):
            pending.append(line)

    if current:
        items.append(current)
    if items:
        return items
    return parse_known_text_pdf(lines, path, supplier, invoice_no)


def read_offer(path: Path, supplier: str | None = None) -> list[SupplierItem]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_supplier_xlsx(path, supplier)
    if suffix == ".pdf":
        return parse_layout_pdf(path, supplier)
    if suffix == ".xls":
        return read_supplier_xls(path, supplier)
    raise ValueError(f"{path.name}: неподдержанный формат КП")


def candidate_request_indexes(offer: SupplierItem, request_items: list[RequestItem], token_index: dict[str, set[int]], num_index: dict[str, set[int]]) -> set[int]:
    offer_tokens = tokens(offer.name)
    offer_nums = numeric_tokens(offer.name)
    candidates: set[int] = set()
    for token in offer_tokens:
        candidates.update(token_index.get(token, set()))
    for number in offer_nums:
        candidates.update(num_index.get(number, set()))

    if not candidates:
        return set(range(len(request_items)))

    req_brands_by_index = {idx: brands(request_items[idx].name) for idx in candidates}
    offer_brands = brands(offer.name)
    if offer_brands:
        brand_matches = {idx for idx, req_brands in req_brands_by_index.items() if req_brands and not req_brands.isdisjoint(offer_brands)}
        if brand_matches:
            candidates = brand_matches

    if len(candidates) < 6:
        number_matches = set()
        for number in offer_nums:
            number_matches.update(num_index.get(number, set()))
        candidates.update(number_matches)
    return candidates or set(range(len(request_items)))


def call_deepseek(payload: dict) -> dict | None:
    api_key = os.environ.get("DEEPSEEK_API_KEY", "").strip()
    if not api_key:
        return None

    body = {
        "model": os.environ.get("DEEPSEEK_MODEL", DEEPSEEK_MODEL),
        "messages": [
            {
                "role": "system",
                "content": (
                    "Ты сопоставляешь строительные материалы из заявки и КП. "
                    "Учитывай синонимы, бренды, размеры, единицы измерения и смысл товара. "
                    "Примеры: метиз/шуруп/саморез/крепеж близкие группы; ГКЛ/гипсокартон близкие; "
                    "лист, рулон и упаковка могут соответствовать м2, если площадь указана в названии. "
                    "Если в названии указано количество в пачке/упаковке, умножай на количество упаковок и сравнивай в штуках. "
                    "Упаковки клея, смеси и сухих материалов сравнивай в кг, если указан вес упаковки. "
                    "Пиломатериалы: доска, брус, рейка, лаги и стропила сравнивай в м3, если указаны размеры. "
                    "Металлический профиль, планки и каркас сравнивай в погонных метрах, если указана длина. "
                    "Не сопоставляй разные бренды, если бренд принципиален. "
                    "Ответь только JSON без пояснений."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "task": (
                            "Для каждой строки offer_positions выбери pos из request_positions или null. "
                            "Если точного совпадения нет, но есть 2-5 вероятных вариантов, верни лучший request_pos и добавь alternatives. "
                            "Верни JSON: {\"matches\":[{\"offer_id\": число, \"request_pos\": строка или null, "
                            "\"confidence\": число от 0 до 1, \"reason\": коротко, "
                            "\"alternatives\":[{\"request_pos\": строка, \"confidence\": число, \"reason\": коротко}]}]}"
                        ),
                        **payload,
                    },
                    ensure_ascii=False,
                ),
            },
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0,
    }
    request = urllib.request.Request(
        deepseek_api_url(),
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    timeout = float(os.environ.get("DEEPSEEK_TIMEOUT", "25"))
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        if exc.code == 402:
            record_ai_warning("DeepSeek не выполнил проверку: на аккаунте нет оплаченного баланса.")
        elif exc.code in {401, 403}:
            record_ai_warning("DeepSeek не выполнил проверку: API-ключ не принят сервисом.")
        else:
            record_ai_warning(f"DeepSeek не выполнил проверку: API вернул ошибку {exc.code}.")
        return None
    except (urllib.error.URLError, TimeoutError, OSError, ValueError):
        record_ai_warning("DeepSeek не выполнил проверку: нет соединения с API или истекло время ожидания.")
        return None

    try:
        content = json.loads(raw)["choices"][0]["message"]["content"]
        return json.loads(content)
    except (KeyError, IndexError, TypeError, json.JSONDecodeError):
        record_ai_warning("DeepSeek ответил в неожиданном формате, проверка ИИ пропущена.")
        return None


ProgressCallback = Callable[[int, int, str], None]


def improve_matches_with_deepseek(
    request_items: list[RequestItem],
    matches: list[Match],
    progress_callback: ProgressCallback | None = None,
) -> list[Match]:
    if not os.environ.get("DEEPSEEK_API_KEY", "").strip():
        return add_low_confidence_review_suggestions(request_items, matches)

    request_payload = [
        {
            "pos": item.pos,
            "name": item.name,
            "specs": item.specs,
            "unit": item.unit,
            "qty": item.qty,
            "area_m2_hint": area_m2_from_text(item.name),
        }
        for item in request_items
    ]
    by_pos = {item.pos: item for item in request_items}
    ai_scope = os.environ.get("DEEPSEEK_AI_SCOPE", "all").strip().lower()
    if ai_scope == "all":
        need_ai = [(idx, match) for idx, match in enumerate(matches) if match.status != "service"]
    else:
        need_ai = [
            (idx, match)
            for idx, match in enumerate(matches)
            if match.status != "service" and (match.status == "review" or not match.request_pos)
        ]
    max_rows = int(os.environ.get("DEEPSEEK_MAX_AI_ROWS", "300"))
    batch_size = int(os.environ.get("DEEPSEEK_BATCH_SIZE", "20"))
    updated = list(matches)
    total_to_check = min(len(need_ai), max_rows)
    if progress_callback:
        progress_callback(0, total_to_check, "ИИ готовит позиции к проверке")
    if len(need_ai) > max_rows:
        record_ai_warning(f"DeepSeek проверил только первые {max_rows} строк КП из {len(need_ai)} по текущему лимиту.")

    for start in range(0, total_to_check, batch_size):
        batch = need_ai[start : start + batch_size]
        offer_payload = [
            {
                "offer_id": idx,
                "supplier": match.supplier_item.supplier,
                "row_no": match.supplier_item.row_no,
                "name": match.supplier_item.name,
                "unit": match.supplier_item.unit,
                "qty": match.supplier_item.qty,
                "area_m2_hint": area_m2_from_text(match.supplier_item.name),
                "price": match.supplier_item.price,
                "current_request_pos": match.request_pos,
                "current_score": round(match.score, 3),
            }
            for idx, match in batch
        ]
        result = call_deepseek({"request_positions": request_payload, "offer_positions": offer_payload})
        done_count = min(start + len(batch), total_to_check)
        if progress_callback:
            progress_callback(done_count, total_to_check, f"ИИ проверил {done_count} из {total_to_check} строк КП")
        if not result:
            continue
        for item in result.get("matches", []):
            try:
                offer_id = int(item.get("offer_id"))
            except (TypeError, ValueError):
                continue
            if offer_id < 0 or offer_id >= len(updated):
                continue
            request_pos = clean_text(item.get("request_pos"))
            confidence = parse_number(item.get("confidence")) or 0
            if confidence > 1:
                confidence /= 100
            alternatives = []
            for alt in item.get("alternatives") or []:
                alt_pos = clean_text(alt.get("request_pos"))
                alt_confidence = parse_number(alt.get("confidence")) or 0
                if alt_confidence > 1:
                    alt_confidence /= 100
                alt_reason = clean_text(alt.get("reason") or "")
                if alt_pos in by_pos:
                    alternatives.append((alt_pos, alt_confidence, alt_reason))
            alternatives = sorted(alternatives, key=lambda alt: alt[1], reverse=True)[:5]
            if (not request_pos or request_pos not in by_pos) and alternatives:
                request_pos, confidence, _ = alternatives[0]
            if not request_pos or request_pos not in by_pos or confidence < 0.22:
                continue
            old = updated[offer_id]
            if product_types_incompatible(by_pos[request_pos].name, old.supplier_item.name):
                updated[offer_id] = Match(old.supplier_item, None, 0, "unmatched", "несовместимые товарные категории")
                continue
            reason = clean_text(item.get("reason") or "")
            alt_text = "; ".join(
                f"{pos} ({round(conf * 100)}%)" for pos, conf, _ in alternatives if pos != request_pos
            )
            if alt_text:
                reason = clean_text(f"{reason}. Другие варианты: {alt_text}")
            if confidence >= 0.72:
                updated[offer_id] = Match(old.supplier_item, request_pos, confidence, "auto", "")
            else:
                updated[offer_id] = Match(
                    old.supplier_item,
                    request_pos,
                    confidence,
                    "review",
                    f"ИИ предлагает проверить: {reason}" if reason else "ИИ предлагает проверить совпадение",
                )
    return add_low_confidence_review_suggestions(request_items, updated)


def add_low_confidence_review_suggestions(
    request_items: list[RequestItem],
    matches: list[Match],
) -> list[Match]:
    """Keep weak candidates visible for manual review instead of dropping them."""
    min_score = float(os.environ.get("MIN_REVIEW_SUGGESTION_SCORE", "0.08"))
    target_percent = float(os.environ.get("REVIEW_SUGGESTION_TARGET_PERCENT", "0.95"))
    target_fill_score = float(os.environ.get("MIN_REVIEW_TARGET_FILL_SCORE", "0.02"))
    updated = list(matches)
    comparable_count = sum(1 for match in matches if match.status != "service")
    matched_count = sum(1 for match in matches if match.status != "service" and match.request_pos)
    target_count = min(comparable_count, math.ceil(comparable_count * max(0, min(1, target_percent))))
    candidates: list[tuple[int, Match, RequestItem, float, str]] = []

    for idx, match in enumerate(matches):
        if match.status == "service" or match.request_pos:
            continue
        best_item = None
        best_score = 0.0
        best_reason = ""
        for request in request_items:
            score, reason = match_score(request.name, match.supplier_item.name)
            if score > best_score:
                best_score = score
                best_item = request
                best_reason = reason
        if best_item and best_score > 0:
            candidates.append((idx, match, best_item, best_score, best_reason))

    for idx, match, best_item, best_score, best_reason in sorted(candidates, key=lambda item: item[3], reverse=True):
        promote_by_score = best_score >= min_score
        promote_to_target = matched_count < target_count and best_score >= target_fill_score
        if not promote_by_score and not promote_to_target:
            continue
        if best_score < min_score:
            reason = best_reason or "очень слабое возможное совпадение, требуется ручная проверка"
        else:
            reason = best_reason or "низкая уверенность сопоставления, требуется ручная проверка"
        updated[idx] = Match(match.supplier_item, best_item.pos, best_score, "review", reason)
        matched_count += 1
    return updated


def build_matches(
    request_items: list[RequestItem],
    supplier_items: list[SupplierItem],
    progress_callback: ProgressCallback | None = None,
) -> list[Match]:
    clear_ai_warnings()
    load_env_file()
    matches: list[Match] = []
    token_index: dict[str, set[int]] = {}
    num_index: dict[str, set[int]] = {}
    for idx, request in enumerate(request_items):
        for token in tokens(request.name):
            token_index.setdefault(token, set()).add(idx)
        for number in numeric_tokens(request.name):
            num_index.setdefault(number, set()).add(idx)

    for offer in supplier_items:
        if is_service_item(offer):
            matches.append(Match(offer, None, 0, "service", "доставка или услуга: учитывается отдельно и не входит в процент сопоставления товаров"))
            continue
        best_item = None
        best_score = 0.0
        best_reason = ""
        for idx in candidate_request_indexes(offer, request_items, token_index, num_index):
            request = request_items[idx]
            score, reason = match_score(request.name, offer.name)
            if score > best_score:
                best_score = score
                best_item = request
                best_reason = reason
        if best_item and best_score >= 0.36:
            status = "auto" if best_score >= 0.62 and not best_reason else "review"
            matches.append(Match(offer, best_item.pos, best_score, status, best_reason))
        elif best_item and (
            best_score >= 0.26
            or (best_score >= 0.20 and not item_categories(best_item.name).isdisjoint(item_categories(offer.name)))
        ):
            matches.append(Match(offer, best_item.pos, best_score, "review", best_reason or "возможное совпадение, требуется проверка"))
        else:
            matches.append(Match(offer, None, best_score, "unmatched", "не найдено надежное совпадение"))
    return improve_matches_with_deepseek(request_items, matches, progress_callback)


def write_review(path: Path, matches: list[Match], request_items: list[RequestItem]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Проверка"
    headers = [
        "Статус",
        "Позиция заявки (исправить при необходимости)",
        "Позиция заявки - текст",
        "Поставщик",
        "Строка КП",
        "Позиция КП",
        "Кол-во КП",
        "Ед.",
        "Цена",
        "Сумма",
        "Общий объем для сравнения",
        "Ед. сравнения",
        "Уверенность",
        "Причина",
        "Источник",
    ]
    ws.append(headers)
    by_pos = {item.pos: item for item in request_items}
    for match in matches:
        request = by_pos.get(match.request_pos or "")
        ws.append(
            [
                status_label(match.status),
                match.request_pos or "",
                request.name if request else "",
                match.supplier_item.supplier,
                match.supplier_item.row_no,
                match.supplier_item.name,
                match.supplier_item.qty,
                match.supplier_item.unit,
                match.supplier_item.price,
                match.supplier_item.total,
                match.supplier_item.override_qty,
                match.supplier_item.override_unit,
                round(match.score, 3),
                match.reason,
                match.supplier_item.source,
            ]
        )
    style_sheet(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [14, 22, 48, 22, 12, 62, 12, 8, 12, 14, 18, 12, 12, 34, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def read_review_overrides(path: Path) -> dict[tuple[str, str, str], str | None]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Проверка"] if "Проверка" in wb.sheetnames else wb.active
    overrides: dict[tuple[str, str, str], str | None] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[3]:
            continue
        request_pos = clean_text(row[1])
        supplier = clean_text(row[3])
        row_no = clean_text(row[4])
        source = clean_text(row[14] if len(row) > 14 else row[12])
        key = (supplier, row_no, source)
        if request_pos.lower() in {"", "-", "нет", "skip", "не сравнивать"}:
            overrides[key] = None
        else:
            overrides[key] = request_pos
    return overrides


def apply_overrides(matches: list[Match], overrides: dict[tuple[str, str, str], str | None]) -> list[Match]:
    result: list[Match] = []
    for match in matches:
        key = (match.supplier_item.supplier, match.supplier_item.row_no, match.supplier_item.source)
        if key in overrides:
            request_pos = overrides[key]
            status = "manual" if request_pos else "unmatched"
            reason = "подтверждено на проверке" if request_pos else "оставлено без сопоставления"
            result.append(Match(match.supplier_item, request_pos, match.score, status, reason))
        else:
            result.append(match)
    return result


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)


def fmt_qty(value: float | None) -> str | float:
    return "" if value is None else value


def delivery_mark(delivery: str) -> str:
    text = clean_text(delivery)
    if not text:
        return "✓"
    if "налич" in text.lower() or "склад" in text.lower():
        return "✓"
    return text


def offer_total_value(offer: SupplierItem) -> float | None:
    if offer.total is not None:
        return offer.total
    if offer.price is not None and offer.qty is not None:
        return offer.price * offer.qty
    return offer.price


def fmt_price_value(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}".rstrip("0").rstrip(".")


def normalized_unit_price(request: RequestItem, offer: SupplierItem) -> tuple[float | None, str | float | None]:
    if offer.price is None:
        return None, offer.price
    if offer.override_qty is not None and offer.override_qty > 0 and offer.override_unit:
        total = offer_total_value(offer)
        if total is not None:
            price = total / offer.override_qty
            display = f"{fmt_price_value(offer.price)}\n~ {fmt_price_value(price)}/{offer.override_unit}".strip()
            return price, display
    converted_qty, converted_unit = converted_quantity(request, offer)
    req_unit = normalized_unit(request.unit)
    offer_unit = normalized_unit(offer.unit)
    display: str | float | None = offer.price
    if converted_unit == "m3" and is_lumber_text(f"{request.name} {offer.name}") and converted_qty is not None and offer.qty and converted_qty > 0:
        factor = converted_qty / offer.qty
        if factor > 0:
            price = offer.price / factor
            if offer_unit != "m3" or abs(factor - 1) > 0.0001:
                display = f"{fmt_price_value(offer.price)}\n~ {fmt_price_value(price)}/м3".strip()
            return price, display
    if (
        converted_qty is not None
        and offer.qty
        and converted_qty > 0
        and converted_unit == req_unit
    ):
        factor = converted_qty / offer.qty
        if factor > 0:
            price = offer.price / factor
            if req_unit != offer_unit or abs(factor - 1) > 0.0001:
                unit_label = request.unit or converted_unit
                display = f"{fmt_price_value(offer.price)}\n~ {fmt_price_value(price)}/{unit_label}".strip()
            return price, display
    return offer.price, display


def write_final(path: Path, request_items: list[RequestItem], matches: list[Match]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"

    suppliers = sorted({m.supplier_item.supplier for m in matches})
    by_request_supplier: dict[tuple[str, str], list[Match]] = {}
    unmatched: list[SupplierItem] = []
    review_items: list[Match] = []
    service_items: list[SupplierItem] = []
    service_by_supplier: dict[str, list[SupplierItem]] = {}
    for match in matches:
        if match.request_pos:
            by_request_supplier.setdefault((match.request_pos, match.supplier_item.supplier), []).append(match)
            if match.status == "review":
                review_items.append(match)
        elif match.status == "service":
            service_items.append(match.supplier_item)
            service_by_supplier.setdefault(match.supplier_item.supplier, []).append(match.supplier_item)
        else:
            unmatched.append(match.supplier_item)

    left_header = PatternFill("solid", fgColor="9FB6C9")
    supplier_fills = [
        PatternFill("solid", fgColor="D9D9D9"),
        PatternFill("solid", fgColor="B6D7A8"),
        PatternFill("solid", fgColor="C9DAF8"),
        PatternFill("solid", fgColor="FCE5CD"),
        PatternFill("solid", fgColor="D9EAD3"),
    ]
    supplier_light_fills = [
        PatternFill("solid", fgColor="EFEFEF"),
        PatternFill("solid", fgColor="E2F0D9"),
        PatternFill("solid", fgColor="EAF2FF"),
        PatternFill("solid", fgColor="FFF2E5"),
        PatternFill("solid", fgColor="EEF7E8"),
    ]
    green = PatternFill("solid", fgColor="C6EFCE")
    dark_green = PatternFill("solid", fgColor="548235")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    gray = PatternFill("solid", fgColor="F2F2F2")
    white = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="8D99A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_range(row_from: int, row_to: int, col_from: int, col_to: int, fill: PatternFill | None = None) -> None:
        for sheet_row in ws.iter_rows(min_row=row_from, max_row=row_to, min_col=col_from, max_col=col_to):
            for cell in sheet_row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if fill:
                    cell.fill = fill

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    ws.cell(1, 1, "позиция\nв заявке")
    ws.cell(1, 2, "Описание закупаемой позиции")
    ws.cell(1, 3, "кол-во")
    style_range(1, 2, 1, 3, left_header)
    for col_idx in range(1, 4):
        ws.cell(1, col_idx).font = Font(bold=True, color="1F2933")

    col = 4
    supplier_start_cols: dict[str, int] = {}
    supplier_fill_by_name: dict[str, PatternFill] = {}
    supplier_light_by_name: dict[str, PatternFill] = {}
    for idx, supplier in enumerate(suppliers):
        supplier_start_cols[supplier] = col
        header_fill = supplier_fills[idx % len(supplier_fills)]
        light_fill = supplier_light_fills[idx % len(supplier_light_fills)]
        supplier_fill_by_name[supplier] = header_fill
        supplier_light_by_name[supplier] = light_fill
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col + 4)
        invoices = supplier_invoice_summary(supplier, matches)
        header_text = supplier
        if invoices:
            header_text = f"{supplier}\nсчет/КП: {invoices}"
        ws.cell(1, col, header_text)
        ws.cell(1, col).font = Font(bold=True, color="1F2933", size=10)
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        style_range(1, 2, col, col + 4, header_fill)
        col += 5

    ws.row_dimensions[1].height = 52
    ws.row_dimensions[2].height = 42
    row = 3
    supplier_goods_totals = {supplier: 0.0 for supplier in suppliers}
    supplier_service_totals = {supplier: sum(offer_total_value(offer) or 0 for offer in service_by_supplier.get(supplier, [])) for supplier in suppliers}
    selected_match_ids: set[int] = set()
    for request in request_items:
        desc_row = row
        label_row = row + 1
        value_row = row + 2

        ws.cell(label_row, 1, "позиция\nв заявке")
        ws.cell(value_row, 1, request.pos)
        ws.merge_cells(start_row=desc_row, start_column=2, end_row=value_row, end_column=2)
        ws.cell(desc_row, 2, request.name)
        ws.merge_cells(start_row=desc_row, start_column=3, end_row=value_row, end_column=3)
        qty_text = fmt_qty(request.qty)
        ws.cell(desc_row, 3, f"{qty_text} {request.unit}".strip())
        style_range(desc_row, value_row, 1, 3, white)
        ws.cell(label_row, 1).font = Font(size=7)
        ws.cell(value_row, 1).font = Font(size=10)
        ws.cell(desc_row, 2).font = Font(bold=True)
        ws.cell(desc_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(desc_row, 3).font = Font(bold=True)

        prices: list[tuple[str, float]] = []
        selected: dict[str, Match] = {}
        for supplier in suppliers:
            candidates = by_request_supplier.get((request.pos, supplier), [])
            if candidates:
                chosen = sorted(
                    candidates,
                    key=lambda item: (item.status != "manual", -(item.supplier_item.price or 0), -item.score),
                )[0]
                selected[supplier] = chosen
                selected_match_ids.add(id(chosen))
                normalized_price, _ = normalized_unit_price(request, chosen.supplier_item)
                if normalized_price is not None:
                    prices.append((supplier, normalized_price))
                supplier_goods_totals[supplier] += offer_total_value(chosen.supplier_item) or 0

        min_price = min((price for _, price in prices), default=None)
        max_price = max((price for _, price in prices), default=None)

        for supplier in suppliers:
            start = supplier_start_cols[supplier]
            match = selected.get(supplier)
            style_range(desc_row, value_row, start, start + 4, supplier_light_by_name[supplier])
            if not match:
                ws.merge_cells(start_row=desc_row, start_column=start, end_row=value_row, end_column=start + 4)
                cell = ws.cell(desc_row, start, "нет в КП")
                cell.font = Font(italic=True, color="667085", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = gray
                continue
            subheaders = ["позиция\nв счете", "цена", "кол-во", "стоимость", "срок"]
            for offset, header in enumerate(subheaders):
                ws.cell(label_row, start + offset, header)
                ws.cell(label_row, start + offset).font = Font(bold=True, size=9)
            offer = match.supplier_item
            ws.merge_cells(start_row=desc_row, start_column=start, end_row=desc_row, end_column=start + 4)
            ws.cell(desc_row, start, offer.name)
            ws.cell(value_row, start, offer.row_no)
            normalized_price, price_display = normalized_unit_price(request, offer)
            ws.cell(value_row, start + 1, price_display)
            qty_result = quantity_check(request, offer)
            ws.cell(value_row, start + 2, qty_result.display)
            ws.cell(value_row, start + 3, offer.total)
            ws.cell(value_row, start + 4, delivery_mark(offer.delivery))
            qty_ok = qty_result.status == "ok"
            if normalized_price is not None and min_price is not None and abs(normalized_price - min_price) < 0.0001:
                ws.cell(value_row, start + 1).fill = green
            if normalized_price is not None and max_price is not None and abs(normalized_price - max_price) < 0.0001 and max_price != min_price:
                ws.cell(value_row, start + 1).fill = red
            qty_cell = ws.cell(value_row, start + 2)
            if qty_result.status == "ok":
                qty_cell.fill = green
            elif qty_result.status == "low":
                qty_cell.fill = yellow
            elif qty_result.status == "high":
                qty_cell.fill = dark_green
                qty_cell.font = Font(bold=True, color="FFFFFF")
            if ws.cell(value_row, start + 4).value == "✓":
                ws.cell(value_row, start + 4).fill = green
            needs_review = match.status == "review" or "разные бренды" in (match.reason or "")
            if needs_review:
                ws.cell(desc_row, start).fill = red if "разные бренды" in (match.reason or "") else yellow
                if qty_result.status == "unknown":
                    ws.cell(value_row, start + 2).fill = yellow
        ws.row_dimensions[desc_row].height = 28
        ws.row_dimensions[label_row].height = 21
        ws.row_dimensions[value_row].height = 25
        row += 3

    for service_index, offer in enumerate(service_items, start=1):
        desc_row = row
        label_row = row + 1
        value_row = row + 2

        ws.cell(label_row, 1, "позиция\nв заявке")
        ws.cell(value_row, 1, f"Д{service_index}")
        ws.merge_cells(start_row=desc_row, start_column=2, end_row=value_row, end_column=2)
        ws.cell(desc_row, 2, offer.name)
        ws.merge_cells(start_row=desc_row, start_column=3, end_row=value_row, end_column=3)
        service_qty = fmt_qty(offer.qty)
        ws.cell(desc_row, 3, f"{service_qty} {offer.unit}".strip())
        style_range(desc_row, value_row, 1, 3, white)
        ws.cell(label_row, 1).font = Font(size=7)
        ws.cell(value_row, 1).font = Font(size=10)
        ws.cell(desc_row, 2).font = Font(bold=True, color="7A3416")
        ws.cell(desc_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(desc_row, 3).font = Font(bold=True)

        for supplier in suppliers:
            start = supplier_start_cols[supplier]
            style_range(desc_row, value_row, start, start + 4, supplier_light_by_name[supplier])
            if supplier != offer.supplier:
                ws.merge_cells(start_row=desc_row, start_column=start, end_row=value_row, end_column=start + 4)
                cell = ws.cell(desc_row, start, "")
                cell.fill = gray
                continue

            subheaders = ["позиция\nв счете", "цена", "кол-во", "стоимость", "срок"]
            for offset, header in enumerate(subheaders):
                ws.cell(label_row, start + offset, header)
                ws.cell(label_row, start + offset).font = Font(bold=True, size=9)
            ws.merge_cells(start_row=desc_row, start_column=start, end_row=desc_row, end_column=start + 4)
            ws.cell(desc_row, start, offer.name)
            ws.cell(value_row, start, offer.row_no)
            ws.cell(value_row, start + 1, offer.price)
            ws.cell(value_row, start + 2, f"{service_qty} {offer.unit}".strip())
            ws.cell(value_row, start + 3, offer_total_value(offer))
            ws.cell(value_row, start + 4, delivery_mark(offer.delivery))
            ws.cell(desc_row, start).fill = yellow
            ws.cell(value_row, start + 3).fill = yellow
            if ws.cell(value_row, start + 4).value == "✓":
                ws.cell(value_row, start + 4).fill = green

        ws.row_dimensions[desc_row].height = 28
        ws.row_dimensions[label_row].height = 21
        ws.row_dimensions[value_row].height = 25
        row += 3

    total_rows = [
        ("Итого по товарам", supplier_goods_totals),
        ("Доставка/услуги", supplier_service_totals),
        (
            "Итого с доставкой",
            {supplier: supplier_goods_totals[supplier] + supplier_service_totals[supplier] for supplier in suppliers},
        ),
    ]
    for label, totals in total_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row, 1, label)
        ws.cell(row, 1).font = Font(bold=True, color="1F2933")
        ws.cell(row, 1).fill = left_header
        style_range(row, row, 1, 3, left_header)
        for supplier in suppliers:
            start = supplier_start_cols[supplier]
            style_range(row, row, start, start + 4, supplier_light_by_name[supplier])
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=start + 2)
            ws.cell(row, start, label)
            ws.cell(row, start).font = Font(bold=True)
            ws.cell(row, start + 3, totals.get(supplier, 0))
            ws.cell(row, start + 3).font = Font(bold=True)
            ws.cell(row, start + 3).number_format = '#,##0.00 ₽'
            if label == "Итого с доставкой":
                ws.cell(row, start).fill = green
                ws.cell(row, start + 3).fill = green
        ws.row_dimensions[row].height = 24
        row += 1

    for sheet_row in ws.iter_rows():
        for cell in sheet_row:
            cell.border = border
            if cell.alignment is None or cell.alignment == Alignment():
                cell.alignment = Alignment(wrap_text=True, vertical="center")

    widths = {1: 9, 2: 52, 3: 12}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(4, ws.max_column + 1):
        offset = (col_idx - 4) % 5
        width = [9, 13, 12, 14, 16][offset]
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for price_col in range(5, ws.max_column + 1, 5):
        for price_cell in ws.iter_cols(min_col=price_col, max_col=price_col, min_row=3, max_row=ws.max_row):
            for cell in price_cell:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00 ₽'
    for total_col in range(7, ws.max_column + 1, 5):
        for total_cell in ws.iter_cols(min_col=total_col, max_col=total_col, min_row=3, max_row=ws.max_row):
            for cell in total_cell:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00 ₽'
    ws.freeze_panes = "D3"

    all_sheet = wb.create_sheet("Все позиции КП")
    headers = [
        "Поставщик",
        "Счет/КП",
        "Строка",
        "Позиция КП",
        "Статус",
        "Позиция заявки",
        "Наименование заявки",
        "В основной сводке",
        "Кол-во",
        "Ед.",
        "Цена",
        "Сумма",
        "Источник",
        "Причина",
    ]
    all_sheet.append(headers)
    request_by_pos = {item.pos: item for item in request_items}
    for match in matches:
        offer = match.supplier_item
        request = request_by_pos.get(match.request_pos or "")
        all_sheet.append(
            [
                offer.supplier,
                invoice_label(offer),
                offer.row_no,
                offer.name,
                status_label(match.status),
                request.pos if request else "",
                request.name if request else "",
                "да" if id(match) in selected_match_ids else "нет",
                offer.qty,
                offer.unit,
                offer.price,
                offer.total,
                offer.source,
                match.reason,
            ]
        )
    style_sheet(all_sheet)
    widths = {"A": 20, "B": 20, "C": 10, "D": 70, "E": 18, "F": 14, "G": 70, "H": 16, "I": 12, "J": 10, "K": 14, "L": 14, "M": 24, "N": 42}
    for col_letter, width in widths.items():
        all_sheet.column_dimensions[col_letter].width = width
    all_sheet.freeze_panes = "A2"
    for col_letter in ("K", "L"):
        for cell in all_sheet[col_letter][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 ₽'

    if unmatched:
        extra = wb.create_sheet("Не сопоставлено")
        headers = ["Поставщик", "Счет/КП", "Строка", "Позиция КП", "Кол-во", "Ед.", "Цена", "Сумма", "Источник"]
        extra.append(headers)
        for offer in unmatched:
            extra.append([offer.supplier, invoice_label(offer), offer.row_no, offer.name, offer.qty, offer.unit, offer.price, offer.total, offer.source])
        style_sheet(extra)
        extra.column_dimensions["A"].width = 20
        extra.column_dimensions["B"].width = 20
        extra.column_dimensions["C"].width = 10
        extra.column_dimensions["D"].width = 70
        extra.column_dimensions["E"].width = 12
        extra.column_dimensions["F"].width = 10
        extra.column_dimensions["G"].width = 14
        extra.column_dimensions["H"].width = 14
        extra.column_dimensions["I"].width = 24
        extra.freeze_panes = "A2"
        for col_letter in ("G", "H"):
            for cell in extra[col_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00 ₽'

    if review_items:
        review_sheet = wb.create_sheet("К проверке")
        headers = ["Поставщик", "Счет/КП", "Строка", "Позиция КП", "Предложенная позиция заявки", "Сходство", "Причина", "Кол-во", "Ед.", "Цена", "Сумма", "Источник"]
        review_sheet.append(headers)
        request_by_pos = {item.pos: item for item in request_items}
        for match in review_items:
            offer = match.supplier_item
            request = request_by_pos.get(match.request_pos or "")
            request_label = f"{request.pos} - {request.name}" if request else ""
            review_sheet.append([
                offer.supplier,
                invoice_label(offer),
                offer.row_no,
                offer.name,
                request_label,
                match.score,
                match.reason,
                offer.qty,
                offer.unit,
                offer.price,
                offer.total,
                offer.source,
            ])
        style_sheet(review_sheet)
        widths = {"A": 20, "B": 20, "C": 10, "D": 70, "E": 70, "F": 12, "G": 40, "H": 12, "I": 10, "J": 14, "K": 14, "L": 24}
        for col_letter, width in widths.items():
            review_sheet.column_dimensions[col_letter].width = width
        review_sheet.freeze_panes = "A2"
        for cell in review_sheet["F"][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'
        for col_letter in ("J", "K"):
            for cell in review_sheet[col_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00 ₽'

    if service_items:
        service = wb.create_sheet("Доставка и услуги")
        headers = ["Поставщик", "Счет/КП", "Строка", "Позиция КП", "Кол-во", "Ед.", "Цена", "Сумма", "Источник"]
        service.append(headers)
        for offer in service_items:
            service.append([offer.supplier, invoice_label(offer), offer.row_no, offer.name, offer.qty, offer.unit, offer.price, offer.total, offer.source])
        style_sheet(service)
        service.column_dimensions["A"].width = 20
        service.column_dimensions["B"].width = 20
        service.column_dimensions["C"].width = 10
        service.column_dimensions["D"].width = 70
        service.column_dimensions["E"].width = 12
        service.column_dimensions["F"].width = 10
        service.column_dimensions["G"].width = 14
        service.column_dimensions["H"].width = 14
        service.column_dimensions["I"].width = 24
        service.freeze_panes = "A2"
        for col_letter in ("G", "H"):
            for cell in service[col_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00 ₽'

    wb.save(path)


def dump_json(path: Path, request_items: list[RequestItem], supplier_items: list[SupplierItem], matches: list[Match]) -> None:
    data = {
        "request_items": [asdict(item) for item in request_items],
        "supplier_items": [asdict(item) for item in supplier_items],
        "matches": [
            {
                **asdict(match),
                "supplier_item": asdict(match.supplier_item),
            }
            for match in matches
        ],
    }
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Сравнение заявки/ТЗ с КП поставщиков и выпуск Excel-сводки.")
    parser.add_argument("--request", required=True, type=Path, help="Файл заявки .xlsx")
    parser.add_argument("--offers", required=True, nargs="+", type=Path, help="Файлы КП: .xlsx или текстовые .pdf")
    parser.add_argument("--out", required=True, type=Path, help="Финальный Excel-файл")
    parser.add_argument("--review", type=Path, help="Файл ручной проверки, который нужно создать")
    parser.add_argument("--review-in", type=Path, help="Заполненный файл ручной проверки для применения правок")
    parser.add_argument("--debug-json", type=Path, help="Сохранить извлеченные данные и сопоставления в JSON")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    request_items = read_request_xlsx(args.request)
    supplier_items: list[SupplierItem] = []
    errors = []
    for offer_path in args.offers:
        try:
            supplier_items.extend(read_offer(offer_path))
        except Exception as exc:  # noqa: BLE001 - CLI should report all bad input files together.
            errors.append(str(exc))
    if errors:
        print("Предупреждения по входным файлам:")
        for error in errors:
            print(f"- {error}")
    if not supplier_items:
        raise SystemExit("Не удалось извлечь ни одной позиции КП.")

    matches = build_matches(request_items, supplier_items)
    ai_warnings = get_ai_warnings()
    if ai_warnings:
        print("Предупреждения по ИИ:")
        for warning in ai_warnings:
            print(f"- {warning}")
    if args.review_in:
        matches = apply_overrides(matches, read_review_overrides(args.review_in))
    if args.review:
        args.review.parent.mkdir(parents=True, exist_ok=True)
        write_review(args.review, matches, request_items)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    write_final(args.out, request_items, matches)
    if args.debug_json:
        args.debug_json.parent.mkdir(parents=True, exist_ok=True)
        dump_json(args.debug_json, request_items, supplier_items, matches)

    print(f"Позиции заявки: {len(request_items)}")
    print(f"Позиции КП: {len(supplier_items)}")
    print(f"Сопоставлено автоматически/на проверку: {sum(1 for m in matches if m.request_pos)}")
    print(f"Несопоставлено: {sum(1 for m in matches if not m.request_pos)}")
    if args.review:
        print(f"Файл проверки: {args.review}")
    print(f"Финальный отчет: {args.out}")


if __name__ == "__main__":
    main()
